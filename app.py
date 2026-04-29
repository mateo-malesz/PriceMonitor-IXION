from flask import Flask, render_template, request, redirect, url_for, flash, session
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from urllib.parse import urlparse
from flask_mail import Mail, Message
from flask import make_response
import xml.etree.ElementTree as ET
from flask_apscheduler import APScheduler
from datetime import datetime, date, timezone
from scraper import get_current_price, init_batch_session, close_batch_session
from flask_admin import Admin, AdminIndexView, expose
from flask_admin.contrib.sqla import ModelView
from sote_integration import fetch_sales_for_date
from datetime import timedelta
import requests
import os
import re
import json
import csv
import io
import logging
from logging.handlers import RotatingFileHandler
from sqlalchemy import func, case
from dotenv import load_dotenv
from threading import Thread
from flask import current_app
try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo


load_dotenv()

# --- KONFIGURACJA STREFY CZASOWEJ ---
TIMEZONE = ZoneInfo("Europe/Warsaw")

def get_current_time():
    return datetime.now(TIMEZONE)

# --- KONFIGURACJA LOGOWANIA ---
# Ustawiamy RotatingFileHandler: max 1MB na plik, trzymamy 5 ostatnich plików
file_handler = RotatingFileHandler('app.log', maxBytes=1024 * 1024, backupCount=5)
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

logging.basicConfig(
    level=logging.INFO,
    handlers=[
        file_handler,
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
# --- KONFIGURACJA MAIL ---
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_RECIPIENT'] = os.getenv('MAIL_RECIPIENT')

mail = Mail(app)


# --- KONFIGURACJA HARMONOGRAMU ---
class Config:
    SCHEDULER_API_ENABLED = True


app.config.from_object(Config())

scheduler = APScheduler()
scheduler.init_app(app)
if os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or not app.debug:
    scheduler.start()

# --- KONFIGURACJA APLIKACJI ---
app.secret_key = os.getenv('SECRET_KEY')
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'database.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "connect_args": {
        "timeout": 30,
    }
}

db = SQLAlchemy(app)

from sqlalchemy import event
from sqlalchemy.engine import Engine

@event.listens_for(Engine, "connect")
def set_sqlite_pragma(dbapi_connection, connection_record):
    if app.config['SQLALCHEMY_DATABASE_URI'].startswith('sqlite'):
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA journal_mode=WAL")
        cursor.execute("PRAGMA synchronous=NORMAL")
        cursor.close()

# --- FLASK-LOGIN ---
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'


# --- FLASK-ADMIN ---
class MyModelView(ModelView):
    def is_accessible(self):
        admin_email = os.getenv('ADMIN_EMAIL')
        return current_user.is_authenticated and admin_email and current_user.email == admin_email

    def inaccessible_callback(self, name, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        flash('Brak uprawnień administratora.', category='error')
        return redirect(url_for('home'))

class ProductModelView(MyModelView):
    column_searchable_list = ['title', 'sku', 'gtin']
    column_filters = ['brand', 'project', 'is_active', 'availability']
    column_list = ['title', 'sku', 'brand', 'project', 'my_price', 'is_active']
    column_sortable_list = ['title', 'sku', 'my_price']

class UserModelView(MyModelView):
    def on_model_change(self, form, model, is_created):
        if form.password.data:
            # SPRAWDZENIE: Jeśli hasło NIE zaczyna się od 'pbkdf2:',
            # to znaczy, że wpisałeś nowe, czyste hasło i trzeba je zahaszować.
            if not form.password.data.startswith('pbkdf2:sha256'):
                model.password = generate_password_hash(form.password.data, method='pbkdf2:sha256')
            else:
                # Jeśli zaczyna się od pbkdf2, to znaczy, że to stary hash
                # – nie dotykamy go, zostawiamy tak jak jest w modelu.
                pass

        return super(UserModelView, self).on_model_change(form, model, is_created)

class ProjectModelView(MyModelView):
    # Wykluczamy ciężkie relacje z formularza (żeby nie wysyłać gigantycznych paczek danych)
    form_excluded_columns = ['products', 'tasks']

    # Opcjonalnie: ułatwiamy sobie widok tabeli, żeby był bardziej przejrzysty
    column_list = ['name', 'domain', 'api_type', 'api_url']

class MyAdminIndexView(AdminIndexView):
    def is_accessible(self):
        admin_email = os.getenv('ADMIN_EMAIL')
        return current_user.is_authenticated and admin_email and current_user.email == admin_email

    def inaccessible_callback(self, name, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        flash('Brak uprawnień administratora.', category='error')
        return redirect(url_for('home'))

    @expose('/')
    def index(self):
        user_count = User.query.count()
        project_count = Project.query.count()
        product_count = Product.query.count()
        return self.render('admin/index.html',
                           user_count=user_count,
                           project_count=project_count,
                           product_count=product_count)


admin = Admin(app, name='Panel Administratora', index_view=MyAdminIndexView())

# --- MODELE BAZY DANYCH ---
project_users = db.Table('project_users',
                         db.Column('user_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
                         db.Column('project_id', db.Integer, db.ForeignKey('project.id'), primary_key=True)
                         )


# --- USER ---
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)
    image_file = db.Column(db.String(500), nullable=False,
                           default='https://ui-avatars.com/api/?name=User&background=0d6efd&color=fff')
    projects = db.relationship('Project', secondary=project_users, backref=db.backref('users', lazy='dynamic'))


# --- PROJEKT ---
class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    domain = db.Column(db.String(100))
    product_feed_url = db.Column(db.String(500), nullable=True)
    last_feed_sync = db.Column(db.DateTime, nullable=True)
    api_type = db.Column(db.String(50), nullable=True)  # np. 'SOTE'
    api_url = db.Column(db.String(500), nullable=True)
    api_user = db.Column(db.String(100), nullable=True)
    api_password = db.Column(db.String(255), nullable=True)
    products = db.relationship('Product', backref='project', lazy=True, cascade="all, delete")


# --- MARKA ---
class Brand(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    products = db.relationship('Product', backref='brand', lazy=True)


# --- SKLEP ---
class Shop(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    domain = db.Column(db.String(100), nullable=False)
    mappings = db.relationship('ProductMapping', backref='shop', lazy=True)


# --- PRODUKT ---
class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    sku = db.Column(db.String(50))
    title = db.Column(db.String(200), nullable=False)
    my_price = db.Column(db.Float)
    my_url = db.Column(db.String(500))
    image_link = db.Column(db.String(500))
    gtin = db.Column(db.String(20))
    availability = db.Column(db.String(20), nullable=True)
    brand_id = db.Column(db.Integer, db.ForeignKey('brand.id'), nullable=True)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    purchase_price = db.Column(db.Float, nullable=True)
    strategic_note = db.Column(db.Text, nullable=True)
    comments = db.relationship('ProductComment', backref='product', lazy=True, cascade="all, delete-orphan")
    mappings = db.relationship('ProductMapping', backref='product', lazy=True, cascade="all, delete")

    @property
    def competitor_count(self):
        count = 0
        for mapping in self.mappings:
            is_my_link = False
            if self.my_url and (mapping.url.strip() == self.my_url.strip()):
                is_my_link = True
            if not is_my_link:
                count += 1
        return count

    def __str__(self):
        return f"{self.sku} - {self.title}"

# --- KOMENTARZE PRODUKTU ---
class ProductComment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=get_current_time)
    user = db.relationship('User', backref='comments')

# --- MAPPING ---
class ProductMapping(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    shop_id = db.Column(db.Integer, db.ForeignKey('shop.id'), nullable=False)
    url = db.Column(db.String(500), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    last_checked_at = db.Column(db.DateTime)
    last_price = db.Column(db.Float)
    history = db.relationship('PriceHistory', backref='mapping', lazy=True, cascade="all, delete")
    is_available = db.Column(db.Boolean, default=True)


# --- HISTORIA CEN ---
class PriceHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    mapping_id = db.Column(db.Integer, db.ForeignKey('product_mapping.id'), nullable=False)
    price = db.Column(db.Float, nullable=False)
    availability = db.Column(db.Boolean, default=True)
    scraped_at = db.Column(db.DateTime, default=get_current_time)


# --- ZADANIE ---
class ScheduledTask(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    brand_id = db.Column(db.Integer, db.ForeignKey('brand.id'), nullable=True)
    run_time = db.Column(db.String(5), nullable=False, default="08:00")
    frequency = db.Column(db.String(20), default='daily')
    days_of_week = db.Column(db.String(5), nullable=True)
    last_run_date = db.Column(db.Date, nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    project = db.relationship('Project', backref=db.backref('tasks', cascade="all, delete-orphan"))
    brand = db.relationship('Brand')


# --- HISTORIA SPRZEDAŻY ---
class SalesHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    quantity = db.Column(db.Integer, default=0, nullable=False)
    revenue = db.Column(db.Float, default=0.0, nullable=False)
    product = db.relationship('Product', backref=db.backref('sales_history', lazy=True, cascade="all, delete"))

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


# --- ROUTING I LOGIKA ---
@app.route('/')
@login_required
def home():
    return redirect(url_for('projects'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('projects'))
        else:
            flash('Błędny email lub hasło.', category='error')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# --- OBSŁUGA BŁĘDÓW ---
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404


# --- FUNKCJE IMPORTU ---
def parse_google_merchant_format(root, project_id):
    stats = {'added': 0, 'updated': 0, 'archived': 0, 'error': None}
    try:
        # Usunięty fragment requests.get, bo plik pobiera nam "rozdzielacz" wyżej
        ns = {'g': 'http://base.google.com/ns/1.0'}

        # Pobieramy wszystkie aktywne produkty z bazy dla tego projektu
        existing_products = {p.sku: p for p in Product.query.filter_by(project_id=project_id).all()}
        imported_skus = set()

        items = root.findall('.//item')
        if not items:
            items = root.findall('.//{http://www.w3.org/2005/Atom}entry')
        brands_cache = {}

        for item in items:
            title_node = item.find('title')
            if title_node is None: title_node = item.find('g:title', ns)
            title = title_node.text.strip() if (title_node is not None and title_node.text) else 'Bez nazwy'

            link_node = item.find('link')
            if link_node is None: link_node = item.find('g:link', ns)
            link = link_node.text.strip() if (link_node is not None and link_node.text) else ''

            sku_node = item.find('g:id', ns)
            if sku_node is None: sku_node = item.find('g:mpn', ns)
            sku_val = sku_node.text.strip() if (sku_node is not None and sku_node.text) else None

            price_node = item.find('g:price', ns)
            price_val = 0.0
            if price_node is not None and price_node.text:
                raw_price = price_node.text.upper().replace('PLN', '').replace('ZŁ', '').replace('EUR', '').replace(' ',
                                                                                                                    '').strip()
                try:
                    price_val = float(raw_price)
                except ValueError:
                    price_val = 0.0

            image_node = item.find('g:image_link', ns)
            image_url = image_node.text.strip() if (image_node is not None and image_node.text) else None

            gtin_node = item.find('g:gtin', ns)
            gtin_val = gtin_node.text.strip() if (gtin_node is not None and gtin_node.text) else None

            avail_node = item.find('g:availability', ns)
            availability = avail_node.text.strip() if (avail_node is not None and avail_node.text) else 'unknown'

            brand_node = item.find('g:brand', ns)
            brand_name = brand_node.text.strip() if (brand_node is not None and brand_node.text) else None
            brand_obj = None

            if brand_name:
                if brand_name in brands_cache:
                    brand_obj = brands_cache[brand_name]
                else:
                    brand_obj = Brand.query.filter_by(name=brand_name).first()
                    if not brand_obj:
                        brand_obj = Brand(name=brand_name)
                        db.session.add(brand_obj)
                        db.session.commit()

                    brands_cache[brand_name] = brand_obj

            if sku_val:
                imported_skus.add(sku_val)

                if sku_val in existing_products:
                    # Aktualizacja istniejącego produktu
                    product = existing_products[sku_val]

                    # 1. Cena
                    if product.my_price != price_val:
                        product.my_price = price_val

                    # 2. Marka
                    new_brand_id = brand_obj.id if brand_obj else None
                    if product.brand_id != new_brand_id:
                        product.brand_id = new_brand_id

                    # 3. Dostępność
                    if product.availability != availability:
                        product.availability = availability

                    # 4. Tytuł (nazwa)
                    if product.title != title:
                        product.title = title

                    # 5. Twój link do produktu
                    if product.my_url != link:
                        product.my_url = link

                    # 6. Zdjęcie
                    if product.image_link != image_url:
                        product.image_link = image_url

                    # 7. Kod GTIN/EAN
                    if product.gtin != gtin_val:
                        product.gtin = gtin_val

                    # Upewniamy się, że produkt jest aktywny (jeśli wrócił z archiwum)
                    if not product.is_active:
                        product.is_active = True

                    stats['updated'] += 1
                else:
                    # Dodanie nowego produktu
                    new_product = Product(
                        project_id=project_id,
                        title=title,
                        sku=sku_val,
                        my_price=price_val,
                        my_url=link,
                        image_link=image_url,
                        gtin=gtin_val,
                        brand_id=brand_obj.id if brand_obj else None,
                        availability=availability,
                        is_active=True
                    )
                    db.session.add(new_product)
                    stats['added'] += 1

        # Archiwizacja produktów, których nie ma w pliku
        for sku, product in existing_products.items():
            if sku not in imported_skus and product.is_active:
                product.is_active = False
                stats['archived'] += 1

        # Aktualizacja daty ostatniej synchronizacji
        project = Project.query.get(project_id)
        if project:
            project.last_feed_sync = get_current_time()

        # Tworzymy brakujące mappingi dla własnego sklepu
        all_products = Product.query.filter(
            Product.project_id == project_id,
            Product.my_url != None,
            Product.is_active == True
        ).all()

        for product in all_products:
            clean_url = product.my_url.strip()
            exists = ProductMapping.query.filter_by(product_id=product.id, url=clean_url).first()
            if not exists:
                try:
                    from urllib.parse import urlparse
                    domain = urlparse(clean_url).netloc.replace('www.', '')
                    if not domain:
                        continue
                    shop = Shop.query.filter_by(domain=domain).first()
                    if not shop:
                        shop = Shop(name=domain.capitalize(), domain=domain)
                        db.session.add(shop)
                        db.session.flush()
                    new_mapping = ProductMapping(
                        product_id=product.id,
                        shop_id=shop.id,
                        url=clean_url,
                        is_active=True
                    )
                    db.session.add(new_mapping)
                    stats['added_mappings'] = stats.get('added_mappings', 0) + 1
                    logger.info(f"[MAPPING] Dodano brakujący mapping dla: {product.sku} -> {domain}")
                except Exception as e:
                    logger.warning(f"[MAPPING] Błąd tworzenia mappingu dla {product.sku}: {e}")
                    continue

        db.session.commit()
        logger.info(f"Import zakończony. Statystyki: {stats}")
        return stats
    except Exception as e:
        logger.critical(f"CRITICAL XML ERROR: {e}", exc_info=True)
        stats['error'] = str(e)
        return stats


def parse_iof_format(root, project_id):
    stats = {'added': 0, 'updated': 0, 'archived': 0, 'error': None}

    # Pobieramy istniejące produkty z bazy
    existing_products = {p.sku: p for p in Product.query.filter_by(project_id=project_id).all()}  #
    imported_skus = set()
    brands_cache = {}

    # W IdoSell produkty są w tagu <products><product>
    products = root.findall('.//product')

    for item in products:
        # SKU / ID
        sku_val = item.get('code_on_card')
        if not sku_val:  # Jeśli nie ma code_on_card, spróbuj wziąć id
            sku_val = item.get('id')

        if not sku_val:  # Jeśli brakuje obu, pomiń produkt
            continue

        # Tytuł
        name_node = item.find('.//description/name')
        title = name_node.text.strip() if (name_node is not None and name_node.text) else 'Bez nazwy'

        # Link do produktu
        card_node = item.find('.//card')
        link = card_node.get('url') if card_node is not None else ''

        # Zdjęcie (pierwsze z listy <large>)
        image_node = item.find('.//images/large/image')
        image_url = image_node.get('url') if image_node is not None else None

        # Cena (bierzemy pierwszą napotkaną cenę brutto z tagu <price>)
        price_node = item.find('.//price')
        price_val = 0.0
        if price_node is not None and price_node.get('gross'):
            try:
                price_val = float(price_node.get('gross').replace(',', '.'))
            except ValueError:
                pass

        # Dostępność z tagu <sizes><size available="...">
        availability = 'OUT OF STOCK'  # domyślnie
        size_node = item.find('.//sizes/size')

        if size_node is not None:
            avail_attr = size_node.get('available')

            # IdoSell ma różne statusy (available, on_order, unavailable itp.)
            if avail_attr in ['available', 'on_order', 'in_stock']:
                availability = 'IN STOCK'
            elif avail_attr == 'unavailable':
                availability = 'OUT OF STOCK'
            else:
                # Jeśli wpadnie jakikolwiek inny status, zapisze jego oryginalną nazwę
                availability = avail_attr if avail_attr else 'OUT OF STOCK'

        # Marka
        producer_node = item.find('.//producer')
        brand_name = producer_node.get('name') if producer_node is not None else None
        brand_obj = None

        if brand_name:
            if brand_name in brands_cache:
                brand_obj = brands_cache[brand_name]
            else:
                brand_obj = Brand.query.filter_by(name=brand_name).first()  #
                if not brand_obj:
                    brand_obj = Brand(name=brand_name)  #
                    db.session.add(brand_obj)  # [cite: 1]
                    db.session.commit()  # [cite: 1]
                brands_cache[brand_name] = brand_obj

        # --- LOGIKA ZAPISU DO BAZY ---
        imported_skus.add(sku_val)

        if sku_val in existing_products:
            # Aktualizacja
            product = existing_products[sku_val]
            if product.my_price != price_val: product.my_price = price_val  # [cite: 1]
            new_brand_id = brand_obj.id if brand_obj else None
            if product.brand_id != new_brand_id: product.brand_id = new_brand_id  # [cite: 1]
            if product.availability != availability: product.availability = availability  # [cite: 1]
            if product.title != title: product.title = title  # [cite: 1]
            if product.my_url != link: product.my_url = link  # [cite: 1]
            if product.image_link != image_url: product.image_link = image_url  # [cite: 1]
            if not product.is_active: product.is_active = True  # [cite: 1]
            stats['updated'] += 1
        else:
            # Dodanie
            new_product = Product(  # [cite: 1]
                project_id=project_id, title=title, sku=sku_val, my_price=price_val,  # [cite: 1]
                my_url=link, image_link=image_url, brand_id=brand_obj.id if brand_obj else None,  # [cite: 1]
                availability=availability, is_active=True  # [cite: 1]
            )
            db.session.add(new_product)  # [cite: 1]
            stats['added'] += 1

    # Archiwizacja i Mappingi (identycznie jak w Twojej starej funkcji)
    for sku, product in existing_products.items():
        if sku not in imported_skus and product.is_active:  # [cite: 1]
            product.is_active = False  # [cite: 1]
            stats['archived'] += 1

    # Mappingi
    all_products = Product.query.filter(Product.project_id == project_id, Product.my_url != None,
                                        Product.is_active == True).all()  # [cite: 1]
    for product in all_products:
        clean_url = product.my_url.strip()  # [cite: 1]
        exists = ProductMapping.query.filter_by(product_id=product.id, url=clean_url).first()  # [cite: 1]
        if not exists:
            try:
                from urllib.parse import urlparse  # [cite: 1]
                domain = urlparse(clean_url).netloc.replace('www.', '')  # [cite: 1]
                if not domain: continue
                shop = Shop.query.filter_by(domain=domain).first()  # [cite: 1]
                if not shop:
                    shop = Shop(name=domain.capitalize(), domain=domain)  # [cite: 1]
                    db.session.add(shop)  # [cite: 1]
                    db.session.flush()  # [cite: 1]
                new_mapping = ProductMapping(product_id=product.id, shop_id=shop.id, url=clean_url,
                                             is_active=True)  # [cite: 1]
                db.session.add(new_mapping)  # [cite: 1]
                stats['added_mappings'] = stats.get('added_mappings', 0) + 1
            except Exception as e:
                continue

    project = Project.query.get(project_id)  # [cite: 1]
    if project: project.last_feed_sync = get_current_time()  # [cite: 1]

    db.session.commit()  # [cite: 1]
    return stats

def import_products_from_xml(url, project_id):
    stats = {'added': 0, 'updated': 0, 'archived': 0, 'error': None}
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'} #[cite: 1]
        response = requests.get(url, headers=headers, timeout=30) #[cite: 1]

        if response.status_code != 200: #[cite: 1]
            logger.error(f"Błąd pobierania XML: {response.status_code}") #[cite: 1]
            stats['error'] = f"Błąd HTTP: {response.status_code}" #[cite: 1]
            return stats

        root = ET.fromstring(response.content) #[cite: 1]

        # Tutaj aplikacja decyduje, którego parsera użyć
        if root.tag == 'offer':
            # Główny tag w IdoSell to <offer>
            logger.info("Wykryto format IdoSell (IOF).")
            return parse_iof_format(root, project_id)
        else:
            # Domyślnie zakładamy Google Merchant
            logger.info("Wykryto format Google Merchant Center.")
            return parse_google_merchant_format(root, project_id) # Zmień nazwę starej funkcji na tę!

    except Exception as e:
        logger.critical(f"CRITICAL XML ERROR: {e}", exc_info=True) #[cite: 1]
        stats['error'] = str(e) #[cite: 1]
        return stats

# --- ROUTING PROJEKTÓW ---
@app.route('/projects')
@login_required
def projects():
    user_projects = current_user.projects
    return render_template('project_list.html', projects=user_projects)


@app.route('/project/new', methods=['GET', 'POST'])
@login_required
def create_project():
    if request.method == 'POST':
        name = request.form.get('name')
        domain = request.form.get('domain')
        import_method = request.form.get('import_method')
        feed_url = request.form.get('feed_url')

        if not name:
            flash('Nazwa projektu jest wymagana!', category='error')
        else:
            # Zapisujemy feed_url w bazie, jeśli został podany
            new_project = Project(name=name, domain=domain, product_feed_url=feed_url if feed_url else None)
            new_project.users.append(current_user)

            db.session.add(new_project)
            db.session.commit()

            if import_method == 'url' and feed_url:
                flash('Rozpoczynam import produktów w tle... To może chwilę potrwać.', category='info')
                result = import_products_from_xml(feed_url, new_project.id)
                if not result['error']:
                    flash(f"Sukces! Dodano: {result['added']}, Zaktualizowano: {result['updated']}.",
                          category='success')
                else:
                    flash(f"Błąd importu: {result['error']}", category='warning')

            elif import_method == 'none':
                flash('Projekt utworzony (pusty).', category='success')

            return redirect(url_for('project_dashboard', project_id=new_project.id))

    return render_template('create_project.html')


# --- USUWANIE PROJEKTU ---
@app.route('/project/<int:project_id>/delete', methods=['POST'])
@login_required
def delete_project(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Nie masz uprawnień do usunięcia tego projektu.', category='error')
        return redirect(url_for('projects'))
    db.session.delete(project)
    db.session.commit()
    flash(f'Projekt "{project.name}" został usunięty.', category='success')
    return redirect(url_for('projects'))


@app.route('/project/<int:project_id>/products')
@login_required
def project_dashboard(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))
    session[f'dashboard_url_{project_id}'] = request.full_path
    # Pobieramy parametry filtrowania i sortowania
    search_query = request.args.get('q', '')
    brand_filter = request.args.get('brand', '')
    availability_filter = request.args.get('availability', '')
    filter_type = request.args.get('filter', '')
    show_archived = request.args.get('archived', 'false') == 'true'

    sort_by = request.args.get('sort', 'title')  # domyślnie po tytule
    sort_order = request.args.get('order', 'asc')  # domyślnie rosnąco

    # Budujemy zapytanie
    query = Product.query.filter_by(project_id=project.id)

    # Filtrowanie po statusie aktywności (archiwum vs aktywne)
    if show_archived:
        query = query.filter_by(is_active=False)
    else:
        query = query.filter_by(is_active=True)

    if search_query:
        query = query.filter(
            (Product.title.ilike(f'%{search_query}%')) |
            (Product.sku.ilike(f'%{search_query}%'))
        )

    if brand_filter and brand_filter.isdigit():
        query = query.filter_by(brand_id=int(brand_filter))

    if availability_filter:
        query = query.filter(Product.availability == availability_filter)

    if filter_type == 'errors':
        query = query.join(ProductMapping).filter(
            ProductMapping.is_active == True,
            (ProductMapping.last_price == None) | (ProductMapping.last_price == 0)
        ).distinct()

    # Sortowanie
    if sort_by == 'title':
        if sort_order == 'desc':
            query = query.order_by(Product.title.desc())
        else:
            query = query.order_by(Product.title.asc())
    elif sort_by == 'brand':
        query = query.join(Brand, isouter=True)
        if sort_order == 'desc':
            query = query.order_by(Brand.name.desc())
        else:
            query = query.order_by(Brand.name.asc())
    elif sort_by == 'sku':
        if sort_order == 'desc':
            query = query.order_by(Product.sku.desc())
        else:
            query = query.order_by(Product.sku.asc())
    elif sort_by == 'price':
        if sort_order == 'desc':
            query = query.order_by(Product.my_price.desc())
        else:
            query = query.order_by(Product.my_price.asc())
    elif sort_by == 'status':
        # Sortowanie po liczbie konkurentów (competitor_count)
        # Musimy użyć podzapytania lub zliczenia w zapytaniu głównym

        # Podzapytanie liczące mappingi, które NIE są linkiem własnym
        stmt = db.session.query(func.count(ProductMapping.id)).filter(
            ProductMapping.product_id == Product.id,
            ProductMapping.url != Product.my_url
        ).scalar_subquery()

        if sort_order == 'desc':
            query = query.order_by(stmt.desc())
        else:
            query = query.order_by(stmt.asc())

    elif sort_by == 'price_index':
        from sqlalchemy import or_

        avg_comp_stmt = db.session.query(func.avg(ProductMapping.last_price)).filter(
            ProductMapping.product_id == Product.id,
            ProductMapping.is_active == True,
            ProductMapping.is_available == True,
            ProductMapping.last_price > 0,
            or_(Product.my_url == None, ProductMapping.url != Product.my_url),
            ProductMapping.last_price >= Product.my_price * 0.2,
            ProductMapping.last_price <= Product.my_price * 5
        ).correlate(Product).scalar_subquery()

        pi_expr = (avg_comp_stmt / Product.my_price) * 100

        if sort_order == 'desc':
            query = query.order_by(pi_expr.desc())
        else:
            # Trik wypychający wartości puste (NULL) na sam dół przy sortowaniu rosnącym
            query = query.order_by(pi_expr.is_(None), pi_expr.asc())

    # Paginacja
    page = request.args.get('page', 1, type=int)
    pagination = query.paginate(page=page, per_page=20, error_out=False)
    filtered_products = pagination.items

    # --- WYLICZANIE PRICE INDEX DLA WIDOKU LISTY ---
    for p in filtered_products:
        p.price_index = None
        if p.my_price and p.my_price > 0:
            valid_prices = [
                m.last_price for m in p.mappings
                if m.is_active and m.is_available and m.last_price and m.last_price > 0
                   and (not p.my_url or m.url.strip() != p.my_url.strip())
                   and (p.my_price * 0.2 <= m.last_price <= p.my_price * 5)  # Odrzucanie anomalii
            ]
            if valid_prices:
                avg_comp = sum(valid_prices) / len(valid_prices)
                p.price_index = round((avg_comp / p.my_price) * 100, 1)

    # Statystyki (liczymy tylko dla AKTYWNYCH produktów)
    all_active_products = Product.query.filter_by(project_id=project.id, is_active=True).all()
    available_brands = db.session.query(Brand).join(Product).filter(Product.project_id == project.id,
                                                                    Product.is_active == True).distinct().order_by(
        Brand.name).all()

    # Pobieramy dostępne statusy dostępności
    available_statuses = db.session.query(Product.availability).filter(Product.project_id == project.id,
                                                                       Product.is_active == True).distinct().order_by(
        Product.availability).all()
    available_statuses = [s[0] for s in available_statuses if s[0]]

    stats = {
        'total_products': len(all_active_products),
        'total_mappings': 0,
        'increased': 0,
        'decreased': 0,
        'avail_in_stock': 0,
        'avail_out_of_stock': 0,
        'avail_other': 0,
        'cheapest_count': 0,
        'expensive_count': 0,
        'archived_count': Product.query.filter_by(project_id=project.id, is_active=False).count()
    }

    for p in all_active_products:
        stats['total_mappings'] += p.competitor_count

        status = str(p.availability).lower() if p.availability else ""

        if 'in stock' in status or 'dostępny' in status or 'available' in status:
            stats['avail_in_stock'] += 1
        elif 'out of stock' in status or 'niedostępny' in status:
            stats['avail_out_of_stock'] += 1
        else:
            stats['avail_other'] += 1

        visible_competitors = [m for m in p.mappings if m.is_active and (not p.my_url or m.url != p.my_url)]
        if p.my_price and visible_competitors:
            competitor_prices = [m.last_price for m in visible_competitors if m.last_price]
            if competitor_prices:
                min_market = min(competitor_prices)
                if p.my_price <= min_market:
                    stats['cheapest_count'] += 1
                else:
                    stats['expensive_count'] += 1

    return render_template('products.html',
                           project=project,
                           products=filtered_products,
                           pagination=pagination,
                           stats=stats,
                           available_brands=available_brands,
                           available_statuses=available_statuses,
                           current_filters={
                               'q': search_query,
                               'brand': brand_filter,
                               'availability': availability_filter,
                               'filter': filter_type,
                               'archived': show_archived,
                               'sort': sort_by,
                               'order': sort_order
                           }
                           )


@app.route('/project/<int:project_id>/add-product', methods=['POST'])
@login_required
def add_product(project_id):
    project = Project.query.get_or_404(project_id)

    if current_user not in project.users:
        flash('Nie kombinuj, to nie Twój projekt.', category='error')
        return redirect(url_for('projects'))

    title = request.form.get('title')
    url = request.form.get('url')
    sku = request.form.get('sku')
    price = request.form.get('price')

    if not title:
        flash('Nazwa produktu jest wymagana!', category='error')
    else:
        if price:
            try:
                price = float(price.replace(',', '.'))
            except ValueError:
                price = 0.0

        new_product = Product(
            title=title,
            my_url=url,
            sku=sku,
            my_price=price,
            project_id=project.id
        )

        db.session.add(new_product)
        db.session.commit()
        flash('Produkt dodany pomyślnie!', category='success')

    return redirect(url_for('project_dashboard', project_id=project_id))


@app.route('/project/<int:project_id>/sync', methods=['POST'])
@login_required
def sync_products(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    if not project.product_feed_url:
        flash('Brak skonfigurowanego linku do pliku XML.', category='warning')
        return redirect(url_for('project_dashboard', project_id=project.id))

    flash('Rozpoczynam synchronizację...', category='info')
    try:
        result = import_products_from_xml(project.product_feed_url, project.id)
        if not result['error']:
            flash(f"Synchronizacja zakończona. Dodano: {result['added']}, Zaktualizowano: {result['updated']}, Zarchiwizowano: {result['archived']}, Nowych mappingów: {result.get('added_mappings', 0)}.", category='success')
        else:
            flash(f"Wystąpił błąd: {result['error']}", category='error')
    except Exception as e:
        logger.error(f"Sync error: {e}")
        flash('Wystąpił błąd krytyczny podczas synchronizacji.', category='error')

    return redirect(url_for('project_dashboard', project_id=project.id))


@app.route('/project/<int:project_id>/force-sales-sync', methods=['POST'])
@login_required
def force_sales_sync(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    if project.api_type != 'SOTE' or not project.api_url:
        flash('Ten projekt nie ma skonfigurowanej integracji SOTE.', category='warning')
        return redirect(url_for('project_dashboard', project_id=project.id))

    # 1. SZUKAMY OSTATNIEJ ZAPISANEJ DATY W BAZIE DLA TEGO PROJEKTU
    last_record = SalesHistory.query.join(Product).filter(Product.project_id == project.id).order_by(
        SalesHistory.date.desc()).first()

    today = date.today()
    yesterday = today - timedelta(days=1)

    if last_record:
        # Zaczynamy dzień po ostatnim wpisie
        current_date = last_record.date + timedelta(days=1)
    else:
        # Jeśli baza jest całkowicie pusta, pobieramy np. z ostatnich 7 dni na start
        current_date = today - timedelta(days=7)

    # Jeśli baza jest już zaktualizowana do wczoraj
    if current_date > yesterday:
        flash('Dane są już w 100% aktualne. Brak dni do nadrobienia!', category='info')
        return redirect(url_for('project_dashboard', project_id=project.id))

    days_processed = 0
    active_products = Product.query.filter_by(project_id=project.id, is_active=True).all()

    # 2. PĘTLA CZASU: LECIMY DZIEŃ PO DNIU AŻ DO WCZORAJ
    while current_date <= yesterday:
        logger.info(f"[SYNC] Nadrabianie daty: {current_date}")

        sales_data = fetch_sales_for_date(project.api_url, project.api_user, project.api_password, current_date)

        for p in active_products:
            sku_upper = str(p.sku).strip().upper() if p.sku else None
            if not sku_upper: continue

            qty = int(sales_data.get(sku_upper, {}).get('qty', 0))
            revenue = float(sales_data.get(sku_upper, {}).get('revenue', 0.0))

            # Zabezpieczenie przed dublami
            existing = SalesHistory.query.filter_by(product_id=p.id, date=current_date).first()
            if existing:
                existing.quantity = qty
                existing.revenue = revenue
            else:
                new_history = SalesHistory(product_id=p.id, date=current_date, quantity=qty, revenue=revenue)
                db.session.add(new_history)

        db.session.commit()
        current_date += timedelta(days=1)
        days_processed += 1

    flash(f'Sukces! Nadrobiono zaległości z {days_processed} dni.', category='success')
    return redirect(url_for('project_dashboard', project_id=project.id))

# --- IMPORT LINKÓW KONKURENCJI ---
@app.route('/project/<int:project_id>/import-links', methods=['GET', 'POST'])
@login_required
def import_links(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Nie wybrano pliku.', category='error')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('Nie wybrano pliku.', category='error')
            return redirect(request.url)

        if file:
            try:
                # Odczytujemy plik w pamięci jako bajty
                file_bytes = file.stream.read()

                # Próbujemy zdekodować różnymi kodowaniami
                decoded_file = None
                encodings = ['utf-8', 'windows-1250', 'latin-1']

                for encoding in encodings:
                    try:
                        decoded_file = file_bytes.decode(encoding)
                        break
                    except UnicodeDecodeError:
                        continue

                if decoded_file is None:
                    flash('Nie udało się rozpoznać kodowania pliku. Upewnij się, że to poprawny plik CSV.',
                          category='error')
                    return redirect(request.url)

                stream = io.StringIO(decoded_file, newline=None)

                # Próbujemy wykryć dialekt (separator)
                try:
                    dialect = csv.Sniffer().sniff(stream.read(1024))
                    stream.seek(0)
                except csv.Error:
                    # Domyślny fallback, jeśli sniffer zawiedzie
                    dialect = csv.excel
                    dialect.delimiter = ';'  # Zakładamy średnik jako domyślny w PL
                    stream.seek(0)

                csv_reader = csv.reader(stream, dialect)

                # Pomijamy nagłówek
                try:
                    next(csv_reader)
                except StopIteration:
                    flash('Plik jest pusty.', category='error')
                    return redirect(request.url)

                added_links = 0
                skipped_duplicates = 0
                products_not_found = 0

                # Cache dla sklepów, żeby nie pytać bazy za każdym razem
                shops_cache = {shop.domain: shop for shop in Shop.query.all()}

                for row in csv_reader:
                    if len(row) < 2: continue  # Za mało kolumn

                    # Kolumna B: SKU / Identyfikator
                    sku = row[1].strip()
                    if not sku: continue

                    product = Product.query.filter_by(project_id=project.id, sku=sku).first()

                    if not product:
                        products_not_found += 1
                        continue

                    # Linki konkurencji zaczynają się od kolumny G (indeks 6)
                    competitor_links = row[6:]

                    for link in competitor_links:
                        link = link.strip()
                        if not link: continue

                        # Sprawdzamy duplikaty
                        exists = ProductMapping.query.filter_by(product_id=product.id, url=link).first()
                        if exists:
                            skipped_duplicates += 1
                            continue

                        # Wyciągamy domenę
                        try:
                            parsed_uri = urlparse(link)
                            domain = parsed_uri.netloc.replace('www.', '')
                            if not domain: continue
                        except:
                            continue

                        # Szukamy lub tworzymy sklep
                        shop = shops_cache.get(domain)
                        if not shop:
                            shop = Shop(name=domain.capitalize(), domain=domain)
                            db.session.add(shop)
                            db.session.flush()  # Żeby dostać ID
                            shops_cache[domain] = shop

                        # Dodajemy mapping
                        new_mapping = ProductMapping(
                            product_id=product.id,
                            shop_id=shop.id,
                            url=link,
                            is_active=True
                        )
                        db.session.add(new_mapping)
                        added_links += 1

                db.session.commit()

                msg = f"Import zakończony. Dodano {added_links} linków."
                if skipped_duplicates > 0:
                    msg += f" Pominięto {skipped_duplicates} duplikatów."
                if products_not_found > 0:
                    msg += f" Nie znaleziono {products_not_found} produktów po SKU."

                flash(msg, category='success')
                return redirect(url_for('project_dashboard', project_id=project.id))

            except Exception as e:
                logger.error(f"CSV Import Error: {e}", exc_info=True)
                flash(f'Wystąpił błąd podczas przetwarzania pliku: {e}', category='error')

    return render_template('import_links.html', project=project)

def import_purchase_prices(file_bytes, filename, project_id):
    import re
    logger.info(f"[IMPORT] Start funkcji, filename={filename}")
    stats = {'updated': 0, 'not_found': 0, 'skipped': 0, 'error': None}
    logger.info(f"[IMPORT] filename ends xlsx: {filename.endswith('.xlsx')}, ends csv: {filename.endswith('.csv')}")
    try:
        # Obsługa XLSX i CSV
        if filename.endswith('.xlsx'):
            logger.info("[IMPORT] Ładuję openpyxl...")
            import openpyxl
            logger.info("[IMPORT] openpyxl załadowany, otwieram workbook...")
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            logger.info(f"[IMPORT] Workbook otwarty: {wb.sheetnames}")
            ws = wb.active
            logger.info(f"[IMPORT] Arkusz aktywny: {ws.dimensions}")

            # Szukamy kolumn po nagłówku, nie indeksie
            headers = {}
            for cell in ws[1]:
                if cell.value:
                    headers[str(cell.value).strip()] = cell.column - 1  # 0-based

            sku_col = headers.get('Symbol')
            price_col = next((v for k, v in headers.items() if 'ostatnia cena zakupu brutto' in k.lower() and 'brutto -' in k.lower()), None)

            logger.info(f"[IMPORT] Znalezione nagłówki: {list(headers.keys())}")
            logger.info(f"[IMPORT] sku_col={sku_col}, price_col={price_col}")

            if sku_col is None or price_col is None:
                stats['error'] = f"Nie znaleziono wymaganych kolumn. Znalezione: {list(headers.keys())}"
                logger.error(f"[IMPORT] Brak kolumn! {stats['error']}")
                return stats

            logger.info(f"[IMPORT] Zaczynam przetwarzać wiersze...")
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            logger.info(f"[IMPORT] Wierszy do przetworzenia: {len(rows)}")




        elif filename.endswith('.csv'):

            logger.info("[IMPORT CSV] Zaczynam dekodowanie...")

            decoded = None

            for enc in ['utf-8-sig', 'utf-8', 'windows-1250', 'latin-1']:

                try:

                    decoded = file_bytes.decode(enc)

                    logger.info(f"[IMPORT CSV] Zdekodowano jako {enc}")

                    break

                except UnicodeDecodeError:

                    continue

            if not decoded:
                stats['error'] = "Nie udało się odczytać pliku CSV."

                logger.error("[IMPORT CSV] Żadne dekodowanie nie zadziałało.")

                return stats

            # --- PANCERNY BLOK DO ŁAPANIA UKRYTYCH BŁĘDÓW ---

            try:

                logger.info(f"[IMPORT CSV] Tworzę StringIO, długość tekstu: {len(decoded)}")

                # Zabezpieczenie: importy lokalne, żeby wykluczyć NameError na tym etapie

                import io

                import csv

                import re

                stream = io.StringIO(decoded)

                logger.info("[IMPORT CSV] io.StringIO utworzone. Czytam CSV (csv.reader)...")

                reader = list(csv.reader(stream, delimiter=';'))

                logger.info(f"[IMPORT CSV] Zbudowano listę wierszy. Łącznie wierszy: {len(reader)}")

                if not reader:
                    stats['error'] = "Plik jest pusty."

                    logger.warning("[IMPORT CSV] Przerwano: plik pusty po przetworzeniu.")

                    return stats

                header_row = reader[0]

                headers = {str(h).strip(): i for i, h in enumerate(header_row)}

                logger.info(f"[IMPORT CSV] Nagłówki: {headers}")

                sku_col = headers.get('Symbol')

                price_col = next((v for k, v in headers.items() if

                                  'ostatnia cena zakupu brutto' in k.lower() and 'brutto -' in k.lower()), None)

                logger.info(f"[IMPORT CSV] Zmapowane kolumny: sku_col={sku_col}, price_col={price_col}")

                if sku_col is None or price_col is None:
                    stats['error'] = f"Nie znaleziono wymaganych kolumn. Znalezione: {list(headers.keys())}"

                    logger.error(f"[IMPORT CSV] Brak kolumn. Przerywam.")

                    return stats

                rows = reader[1:]

                logger.info(f"[IMPORT CSV] Przechodzę do aktualizacji {len(rows)} wierszy w bazie.")


            except Exception as e:

                # To wyłapie KAŻDY błąd i wypluje go do konsoli z pełną ścieżką

                logger.error(f"[CRITICAL ERROR CSV] Wywrotka podczas analizy pliku: {str(e)}", exc_info=True)

                stats['error'] = f"Błąd krytyczny serwera: {str(e)}"

                return stats


        else:

            stats['error'] = "Nieobsługiwany format pliku. Użyj .xlsx lub .csv"

            return stats

        # Pobieramy wszystkie produkty projektu do cache (uppercase SKU)
        products_cache = {
            p.sku.strip().upper(): p
            for p in Product.query.filter_by(project_id=project_id).all()
            if p.sku
        }

        for row in rows:
            try:
                sku_raw = row[sku_col]
                price_raw = row[price_col]

                if not sku_raw:
                    stats['skipped'] += 1
                    continue

                sku_normalized = str(sku_raw).strip().upper()
                product = products_cache.get(sku_normalized)

                if not product:
                    stats['not_found'] += 1
                    continue

                if price_raw is None or str(price_raw).strip() == '':
                    stats['skipped'] += 1
                    continue

                # Parsowanie ceny (obsługa przecinka i kropki)
                price_str = str(price_raw).replace(' ', '').replace(',', '.')
                price_str = re.sub(r'[^\d.]', '', price_str)
                purchase_price = float(price_str)

                if purchase_price <= 2:
                    stats['skipped'] += 1
                    continue

                product.purchase_price = purchase_price
                stats['updated'] += 1

            except (ValueError, IndexError):
                stats['skipped'] += 1
                continue

        db.session.commit()
        logger.info(f"Import cen zakupu: {stats}")
        return stats

    except Exception as e:
        logger.error(f"Błąd importu cen zakupu: {e}", exc_info=True)
        stats['error'] = str(e)
        return stats


@app.route('/project/<int:project_id>/import-purchase-prices', methods=['GET', 'POST'])
@login_required
def import_purchase_prices_view(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    if request.method == 'POST':
        logger.info(f"[IMPORT] Otrzymano POST, pliki: {request.files}")

        if 'file' not in request.files:
            flash('Nie wybrano pliku.', category='error')
            return redirect(request.url)

        file = request.files['file']

        if file.filename == '':
            flash('Nie wybrano pliku.', category='error')
            return redirect(request.url)

        filename = file.filename.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.csv')):
            flash('Obsługiwane formaty: .xlsx, .csv', category='error')
            return redirect(request.url)

        logger.info(f"[IMPORT] Plik: {file.filename}, rozmiar: {file.content_length}")

        # CZYTAMY PLIK TYLKO RAZ
        file_bytes = file.stream.read()
        logger.info(f"[IMPORT] Wczytano bajtów: {len(file_bytes)}")

        result = import_purchase_prices(file_bytes, filename, project_id)

        if result['error']:
            flash(f"Błąd importu: {result['error']}", category='error')
        else:
            msg = f"Zaktualizowano ceny zakupu dla {result['updated']} produktów."
            if result['not_found'] > 0:
                msg += f" Nie znaleziono {result['not_found']} SKU."
            if result['skipped'] > 0:
                msg += f" Pominięto {result['skipped']} wierszy."
            flash(msg, category='success')

        return redirect(url_for('project_dashboard', project_id=project_id))

    return render_template('import_purchase_prices.html', project=project)

# --- WIDOK SZCZEGÓŁÓW PRODUKTU ---
@app.route('/project/<int:project_id>/product/<int:product_id>')
@login_required
def product_details(project_id, product_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    product = Product.query.get_or_404(product_id)

    # --- WYLICZANIE PRICE INDEX DLA PRODUKTU ---
    price_index = None
    if product.my_price and product.my_price > 0:
        valid_prices = [
            m.last_price for m in product.mappings
            if m.is_active and m.is_available and m.last_price and m.last_price > 0
               and (not product.my_url or m.url.strip() != product.my_url.strip())
               and (product.my_price * 0.2 <= m.last_price <= product.my_price * 5)  # Odrzucanie anomalii
        ]
        if valid_prices:
            avg_comp = sum(valid_prices) / len(valid_prices)
            price_index = round((avg_comp / product.my_price) * 100, 1)

    if product.my_url:
        clean_my_url = product.my_url.strip()
        existing_mapping = ProductMapping.query.filter_by(product_id=product.id, url=clean_my_url).first()

        if not existing_mapping:
            domain_name = "Moj Sklep"
            try:
                from urllib.parse import urlparse
                domain_name = urlparse(clean_my_url).netloc.replace('www.', '')
            except:
                pass

            shop = Shop.query.filter_by(domain=domain_name).first()
            if not shop:
                shop = Shop(name=domain_name, domain=domain_name)
                db.session.add(shop)
                db.session.commit()

            new_mapping = ProductMapping(product_id=product.id, shop_id=shop.id, url=clean_my_url, is_active=True)
            db.session.add(new_mapping)
            db.session.commit()

    price_datasets = []
    avail_datasets = []

    # --- ZMIANA SORTOWANIA: Cena rosnąco, potem nazwa sklepu ---
    # Usunąłem warunek "m.url != product.my_url", aby Twój sklep też był sortowany po cenie
    sorted_mappings = sorted(product.mappings, key=lambda m: (
        m.last_price if m.last_price else float('inf'),  # 1. Cena rosnąco (brak ceny na końcu)
        m.shop.name  # 2. Nazwa sklepu
    ))

    for mapping in sorted_mappings:
        if mapping.history:
            sorted_history = sorted(mapping.history, key=lambda x: x.scraped_at)

            price_points = [{'x': h.scraped_at.strftime('%Y-%m-%dT%H:%M:%S'), 'y': h.price} for h in sorted_history]

            avail_points = [{'x': h.scraped_at.strftime('%Y-%m-%dT%H:%M:%S'), 'y': 1 if h.availability else 0} for h in
                            sorted_history]

            is_my_store = False
            if product.my_url and (product.my_url.strip() == mapping.url.strip()):
                is_my_store = True

            if is_my_store:
                color = '#198754'
                border_width = 4
                z_order = 10
            else:
                color = f"#{abs(hash(mapping.shop.name)) % 0xFFFFFF:06x}"
                border_width = 2
                z_order = 1

            price_datasets.append({
                'label': mapping.shop.name + (" (Ty)" if is_my_store else ""),
                'data': price_points,
                'borderColor': color,
                'backgroundColor': color,
                'borderWidth': border_width,
                'pointRadius': 0,
                'fill': False,
                'tension': 0.1,
                'order': z_order
            })

            fill_color = color + "80"

            avail_datasets.append({
                'label': mapping.shop.name + (" (Ty)" if is_my_store else ""),
                'data': avail_points,
                'borderColor': color,
                'backgroundColor': color,
                'borderWidth': 2,
                'pointRadius': 0,
                'fill': False,
                'stepped': True,
                'hidden': not is_my_store,
                'order': z_order
            })

    # --- ZMIANA: Przekazujemy wszystkie mappingi, włącznie z własnym sklepem ---
    # Wcześniej filtrowaliśmy własny sklep, teraz go zostawiamy, aby wyświetlić w tabeli

    # --- POBIERANIE DANYCH O SPRZEDAŻY DLA WYKRESU (ostatnie 30 dni) ---
    from datetime import date, timedelta
    thirty_days_ago = date.today() - timedelta(days=30)

    sales_history = SalesHistory.query.filter(
        SalesHistory.product_id == product.id,
        SalesHistory.date >= thirty_days_ago
    ).order_by(SalesHistory.date.asc()).all()

    qty_points = [{'x': s.date.strftime('%Y-%m-%d'), 'y': s.quantity} for s in sales_history]
    rev_points = [{'x': s.date.strftime('%Y-%m-%d'), 'y': s.revenue} for s in sales_history]
    # Szybkie podsumowanie 30-dniowe
    total_qty = sum(s.quantity for s in sales_history)
    total_rev = sum(s.revenue for s in sales_history)
    sales_summary = {'qty': total_qty, 'rev': total_rev}

    sales_datasets = []
    if qty_points:
        # Słupki ze sztukami
        sales_datasets.append({
            'type': 'bar',
            'label': 'Ilość (szt.)',
            'data': qty_points,
            'backgroundColor': 'rgba(13, 110, 253, 0.7)',
            'yAxisID': 'y',
            'order': 2
        })
        # Linia przychodu
        sales_datasets.append({
            'type': 'line',
            'label': 'Przychód (PLN)',
            'data': rev_points,
            'borderColor': 'rgba(25, 135, 84, 1)',
            'backgroundColor': 'rgba(25, 135, 84, 0.1)',
            'borderWidth': 2,
            'fill': True,
            'tension': 0.3,
            'yAxisID': 'y_rev',
            'order': 1
        })

    return render_template('product_details.html',
                           project=project,
                           product=product,
                           price_index=price_index,
                           chart_data=json.dumps(price_datasets),
                           avail_data=json.dumps(avail_datasets),
                           sales_data=json.dumps(sales_datasets),
                           sales_summary=sales_summary,
                           mappings=sorted_mappings)


# --- DODAWANIE LINKU DO ŚLEDZENIA ---
@app.route('/project/<int:project_id>/product/<int:product_id>/add-url', methods=['POST'])
@login_required
def add_competitor_url(project_id, product_id):
    url = request.form.get('url')

    if not url:
        flash('Musisz podać link!', category='error')
    else:
        existing_mapping = ProductMapping.query.filter_by(product_id=product_id, url=url).first()
        if existing_mapping:
            flash('Ten link jest już monitorowany dla tego produktu!', category='warning')
            return redirect(url_for('product_details', project_id=project_id, product_id=product_id))

        parsed_uri = urlparse(url)
        domain = parsed_uri.netloc.replace('www.', '')

        if not domain:
            flash('Niepoprawny link URL.', category='error')
            return redirect(url_for('product_details', project_id=project_id, product_id=product_id))

        shop = Shop.query.filter_by(domain=domain).first()
        if not shop:
            shop = Shop(name=domain.capitalize(), domain=domain)
            db.session.add(shop)
            db.session.commit()

        new_mapping = ProductMapping(
            product_id=product_id,
            shop_id=shop.id,
            url=url
        )

        db.session.add(new_mapping)
        db.session.commit()
        flash(f'Dodano link do monitorowania ({domain})!', category='success')

    return redirect(url_for('product_details', project_id=project_id, product_id=product_id))


@app.route('/project/<int:project_id>/product/<int:product_id>/refresh', methods=['POST'])
@login_required
def refresh_prices(project_id, product_id):
    project = Project.query.get_or_404(project_id)
    product = Product.query.get_or_404(product_id)

    updated_count = 0

    logger.info(f"--- [REFRESH] Odświeżam produkt: {product.title} ---")

    session = init_batch_session()

    for mapping in product.mappings:
        if mapping.is_active:
            logger.info(f" -> Sprawdzam URL: {mapping.url}")

            try:
                result = get_current_price(mapping.url, session)

                if not result or not isinstance(result, tuple):
                    logger.warning(f"    !!! BŁĄD: Scraper zwrócił błędne dane: {result}")
                    continue

                new_price, is_avail = result

                if new_price is not None:
                    mapping.last_price = new_price
                    mapping.is_available = is_avail
                    mapping.last_checked_at = datetime.now()

                    history = PriceHistory(
                        mapping_id=mapping.id,
                        price=new_price,
                        availability=is_avail
                    )
                    db.session.add(history)
                    updated_count += 1
                    logger.info(f"    -> Sukces: {new_price} PLN (Dostępny: {is_avail})")

                    # --- AKTUALIZACJA CENY WŁASNEJ ---
                    if product.my_url and mapping.url.strip() == product.my_url.strip():
                        product.my_price = new_price
                        logger.info(f"    -> Zaktualizowano cenę własną produktu na: {new_price} PLN")
                    db.session.commit()
                else:
                    logger.warning("    -> Brak ceny (strona nie zwróciła wyniku)")

            except Exception as e:
                logger.error(f"    !!! KRTYYCZNY BŁĄD przy linku {mapping.url}: {e}", exc_info=True)
                continue

    close_batch_session(session)
    db.session.commit()

    if updated_count > 0:
        flash(f'Zaktualizowano ceny w {updated_count} sklepach!', category='success')
    else:
        flash('Nie udało się pobrać żadnej nowej ceny (sprawdź konsolę, aby zobaczyć błędy).', category='warning')

    return redirect(url_for('product_details', project_id=project_id, product_id=product_id))


@app.route('/project/<int:project_id>/product/<int:product_id>/restore', methods=['POST'])
@login_required
def restore_product(project_id, product_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    product = Product.query.get_or_404(product_id)
    product.is_active = True
    db.session.commit()

    flash(f'Produkt "{product.title}" został przywrócony.', category='success')
    return redirect(url_for('project_dashboard', project_id=project.id, archived='true'))


# --- USUWANIE PRODUKTU ---
@app.route('/project/<int:project_id>/product/<int:product_id>/delete', methods=['POST'])
@login_required
def delete_product(project_id, product_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    product = Product.query.get_or_404(product_id)

    db.session.delete(product)
    db.session.commit()

    flash('Produkt został usunięty.', category='success')
    return redirect(url_for('project_dashboard', project_id=project.id))


# --- USUWANIE LINKU KONKURENCJI (MAPPINGU) ---
@app.route('/project/<int:project_id>/mapping/<int:mapping_id>/delete', methods=['POST'])
@login_required
def delete_mapping(project_id, mapping_id):
    # Pobieramy mapping
    mapping = ProductMapping.query.get_or_404(mapping_id)

    if mapping.product.project_id != project_id or current_user not in mapping.product.project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    product_id = mapping.product_id

    db.session.delete(mapping)
    db.session.commit()

    flash('Link do konkurencji usunięty.', category='success')
    return redirect(url_for('product_details', project_id=project_id, product_id=product_id))


@app.route('/project/<int:project_id>/schedule/add', methods=['POST'])
@login_required
def add_task(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    # Pobieramy dane z nowego formularza
    brand_id = request.form.get('brand_id')
    run_time = request.form.get('run_time')
    frequency = request.form.get('frequency')
    days = request.form.getlist('days')

    target_brand = None
    if brand_id and brand_id != 'all':
        target_brand = int(brand_id)

    days_str = ",".join(days) if frequency == 'weekly' else None

    new_task = ScheduledTask(
        project_id=project.id,
        brand_id=target_brand,
        run_time=run_time,
        frequency=frequency,
        days_of_week=days_str
    )

    db.session.add(new_task)
    db.session.commit()

    flash('Dodano nowe zadanie do harmonogramu.', 'success')
    return redirect(url_for('project_scheduler', project_id=project.id))


# --- WYŚWIETLANIE HARMONOGRAMU ---
@app.route('/project/<int:project_id>/scheduler', methods=['GET'])
@login_required
def project_scheduler(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    tasks = ScheduledTask.query.filter_by(project_id=project.id).all()
    brands = Brand.query.join(Product).filter(Product.project_id == project.id).distinct().all()

    return render_template('scheduler.html', project=project, tasks=tasks, brands=brands)


@app.route('/project/<int:project_id>/scheduler/<int:task_id>/delete', methods=['POST'])
@login_required
def delete_task(project_id, task_id):
    task = ScheduledTask.query.get_or_404(task_id)

    if task.project_id != project_id:
        flash('Błąd autoryzacji.', category='error')
        return redirect(url_for('projects'))

    db.session.delete(task)
    db.session.commit()
    flash('Zadanie usunięte.', category='success')
    return redirect(url_for('project_scheduler', project_id=project_id))

def send_enhanced_report(task_name, scan_results):
    logger.info(f"[MAIL] Rozpoczynam przygotowanie raportu: {task_name}")
    if not scan_results:
        logger.info("[MAIL] Brak wyników do wysłania.")
        return

    try:
        # --- CSV ---
        csv_buffer = io.StringIO()
        csv_writer = csv.writer(csv_buffer, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        csv_writer.writerow(['Produkt', 'SKU', 'Sklep', 'Status', 'Stara Cena', 'Nowa Cena', 'Zmiana %', 'Link'])
        for item in scan_results:
            diff = ""
            if item.get('old_price') and item.get('new_price') and item['old_price'] > 0:
                diff = f"{round((item['new_price'] - item['old_price']) / item['old_price'] * 100, 1)}%"
            csv_writer.writerow([
                item['product'], item['sku'], item['shop'],
                item['status'].upper(),
                item['old_price'] if item['old_price'] else '',
                item['new_price'] if item['new_price'] else '',
                diff, item['url']
            ])
        csv_buffer.seek(0)

        APP_URL = os.environ.get('APP_URL', 'http://192.168.24.112:5005')

        # --- GRUPOWANIE PO PRODUKCIE ---
        # Zbieramy tylko produkty z aktywnością (zmiany lub błędy)
        products_map = {}
        for item in scan_results:
            if item['status'] == 'ok':
                continue
            key = item.get('product_id', item['product'])
            if key not in products_map:
                products_map[key] = {
                    'name': item['product'],
                    'sku': item['sku'],
                    'product_id': item.get('product_id'),
                    'project_id': item.get('project_id'),
                    'my_price': item.get('my_price'),
                    'rows': []
                }
            products_map[key]['rows'].append(item)

        total_changes = sum(1 for i in scan_results if i['status'] == 'change')
        total_errors = sum(1 for i in scan_results if i['status'] == 'error')

        def pct_val(item):
            if item.get('old_price') and item['old_price'] > 0 and item.get('new_price'):
                return (item['new_price'] - item['old_price']) / item['old_price'] * 100
            return 0

        def price_badge(item):
            if item['status'] == 'error':
                return '<span style="color:#dc3545;font-size:12px;">⚠️ Błąd pobierania</span>'
            pct = pct_val(item)
            if pct == 0 and item['status'] == 'change':
                return f'<span style="font-weight:700;">{item["new_price"]} PLN</span>'
            color = "#dc3545" if pct > 0 else "#198754"
            arrow = "▲" if pct > 0 else "▼"
            sign = "+" if pct > 0 else ""
            old = f"{item['old_price']} PLN" if item.get('old_price') else "—"
            new = f"{item['new_price']} PLN" if item.get('new_price') else "—"
            # różnica kwotowa
            diff_pln = ""
            if item.get('old_price') and item.get('new_price'):
                diff_val = item['new_price'] - item['old_price']
                diff_pln = f"{'+' if diff_val > 0 else ''}{round(diff_val, 2)} PLN"
            badge = f'<span style="background:{color};color:white;padding:1px 6px;border-radius:10px;font-size:11px;font-weight:700;">{arrow} {diff_pln} ({sign}{round(pct, 1)}%)</span>'
            return f'<span style="color:#999;">{old}</span> → <span style="font-weight:700;color:{color};">{new}</span> &nbsp;{badge}'

        def build_product_block(prod):
            product_url = f"{APP_URL}/project/{prod['project_id']}/product/{prod['product_id']}" \
                if prod.get('product_id') and prod.get('project_id') else None

            name_html = f'<a href="{product_url}" style="color:#0d6efd;text-decoration:none;font-weight:700;font-size:14px;">{prod["name"]}</a>' \
                if product_url else f'<span style="font-weight:700;font-size:14px;">{prod["name"]}</span>'

            # cena własna obok SKU
            my_price_html = ""
            if prod.get('my_price'):
                my_price_html = f'<span style="color:#198754;font-size:11px;margin-left:8px;font-weight:600;">Twoja cena: {prod["my_price"]} PLN</span>'

            rows_html = ""
            for row in prod['rows']:
                shop_link = f'<a href="{row["url"]}" style="color:#555;text-decoration:none;">{row["shop"]}</a>'
                rows_html += f"""
                <tr style="border-top:1px solid #f5f5f5;">
                    <td style="padding:8px 12px 8px 24px;font-size:12px;color:#555;width:160px;">{shop_link}</td>
                    <td style="padding:8px 12px;font-size:13px;">{price_badge(row)}</td>
                </tr>"""

            return f"""
            <tr>
                <td colspan="2" style="padding:14px 12px 4px;">
                    {name_html}
                    <span style="color:#aaa;font-size:11px;margin-left:8px;">SKU: {prod['sku'] or '—'}</span>
                    {my_price_html}
                </td>
            </tr>
            {rows_html}
            <tr><td colspan="2" style="padding:4px 0;"></td></tr>"""

        products_blocks = "".join(build_product_block(p) for p in products_map.values())

        html_body = f"""
        <html><body style="font-family:Arial,sans-serif;color:#333;margin:0;padding:0;background:#f4f6f8;">
        <div style="max-width:620px;margin:30px auto;background:white;border-radius:10px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08);">

            <div style="background:#0d6efd;padding:24px 28px;">
                <div style="color:white;font-size:20px;font-weight:700;">📊 Raport Cenowy</div>
                <div style="color:rgba(255,255,255,0.75);font-size:13px;margin-top:4px;">{task_name} &nbsp;·&nbsp; {date.today().strftime('%d.%m.%Y')}</div>
            </div>

            <div style="display:flex;border-bottom:1px solid #eee;">
                <div style="flex:1;padding:16px;text-align:center;border-right:1px solid #eee;">
                    <div style="font-size:26px;font-weight:700;">{len(scan_results)}</div>
                    <div style="font-size:11px;color:#999;text-transform:uppercase;margin-top:2px;">Sprawdzono</div>
                </div>
                <div style="flex:1;padding:16px;text-align:center;border-right:1px solid #eee;">
                    <div style="font-size:26px;font-weight:700;color:#0d6efd;">{len(products_map)}</div>
                    <div style="font-size:11px;color:#999;text-transform:uppercase;margin-top:2px;">Produkty z aktywnością</div>
                </div>
                <div style="flex:1;padding:16px;text-align:center;border-right:1px solid #eee;">
                    <div style="font-size:26px;font-weight:700;color:#198754;">{total_changes}</div>
                    <div style="font-size:11px;color:#999;text-transform:uppercase;margin-top:2px;">Zmiany cen</div>
                </div>
                <div style="flex:1;padding:16px;text-align:center;">
                    <div style="font-size:26px;font-weight:700;color:#dc3545;">{total_errors}</div>
                    <div style="font-size:11px;color:#999;text-transform:uppercase;margin-top:2px;">Błędy</div>
                </div>
            </div>

            <div style="padding:20px 28px;">
        """

        if products_map:
            html_body += f"""
                <table style="width:100%;border-collapse:collapse;">
                    {products_blocks}
                </table>"""
        else:
            html_body += '<div style="text-align:center;padding:30px;color:#999;font-size:14px;">✅ Brak zmian cen — wszystko stabilne.</div>'

        html_body += f"""
            </div>
            <div style="background:#f8f9fa;padding:20px 28px;text-align:center;border-top:1px solid #eee;">
                <p style="margin:0 0 12px;color:#999;font-size:12px;">Pełna lista w załączonym pliku <b>raport_skanowania.csv</b></p>
                <a href="{APP_URL}" style="background:#0d6efd;color:white;padding:10px 24px;text-decoration:none;border-radius:6px;font-size:13px;font-weight:600;">Przejdź do Panelu →</a>
            </div>
        </div>
        </body></html>"""

        recipient = app.config.get('MAIL_RECIPIENT') or app.config['MAIL_DEFAULT_SENDER']
        msg = Message(f"{task_name} - Raport Cenowy", recipients=[recipient])
        msg.html = html_body
        msg.attach("raport_skanowania.csv", "text/csv", csv_buffer.getvalue().encode('utf-8-sig'))
        mail.send(msg)
        logger.info("--- [MAIL] Raport wysłany! ---")

    except Exception as e:
        logger.critical(f"[MAIL CRITICAL ERROR] {e}", exc_info=True)


@scheduler.task('interval', id='main_scanner_job', seconds=60)
def run_scheduled_scans():
    with app.app_context():
        now = datetime.now()
        current_time = now.strftime("%H:%M")
        today_date = date.today()

        # 1. Sprawdzamy, czy trzeba uruchomić automatyczną synchronizację feedów (np. o 04:00)
        if current_time == "06:00":
            logger.info("--- [AUTO SYNC] Rozpoczynam automatyczną synchronizację feedów ---")
            projects_with_feed = Project.query.filter(Project.product_feed_url != None).all()
            for proj in projects_with_feed:
                if proj.product_feed_url:
                    logger.info(f"Synchronizacja projektu: {proj.name}")
                    import_products_from_xml(proj.product_feed_url, proj.id)

        # 2. Standardowe zadania sprawdzania cen
        tasks = ScheduledTask.query.filter_by(is_active=True).all()
        tasks_to_run = []

        for task in tasks:
            if task.last_run_date == today_date:
                continue
            if task.run_time != current_time:
                continue

            if task.frequency == 'weekly' and task.days_of_week:
                allowed_days = [int(d.strip()) for d in task.days_of_week.split(',') if d.strip().isdigit()]
                if now.weekday() not in allowed_days:
                    continue

            tasks_to_run.append(task)

        if not tasks_to_run:
            return

        logger.info(f"--- [SCHEDULER] Uruchamiam {len(tasks_to_run)} zadań ---")
        batch_session = init_batch_session()

        for task in tasks_to_run:
            task_label = f"Marka {task.brand.name}" if task.brand else "Cały Projekt"
            task_name = f"Raport ({current_time}) - {task_label}"
            scan_results = []

            query = Product.query.filter_by(project_id=task.project_id, is_active=True)
            if task.brand_id:
                query = query.filter_by(brand_id=task.brand_id)
            products = query.all()

            for product in products:
                logger.info(f"[SCAN PRODUKT] {product.title} (SKU: {product.sku})")
                for mapping in product.mappings:
                    if mapping.is_active:
                        logger.info(f"  -> [{mapping.shop.name}] {mapping.url}")
                        old_price = mapping.last_price

                        # logger.info(f"Sprawdzam: {mapping.shop.name} -> {mapping.url[:30]}...")

                        try:
                            result = get_current_price(mapping.url, batch_session)
                            if result and isinstance(result, tuple):
                                new_price, is_avail = result
                                logger.info(f"  -> WYNIK: {new_price} PLN | dostępny: {is_avail}")
                            else:
                                new_price = None
                                is_avail = False
                                logger.warning(f"  -> WYNIK: brak ceny")
                        except Exception as e:
                            logger.error(f"Error scanning {mapping.url}: {e}")
                            new_price = None
                            is_avail = False

                        # result_entry = {
                        #     'product': product.title, 'shop': mapping.shop.name,
                        #     'old_price': old_price, 'new_price': new_price,
                        #     'status': 'ok', 'msg': 'OK', 'sku': product.sku, 'url': mapping.url
                        # }

                        result_entry = {
                            'product': product.title, 'shop': mapping.shop.name,
                            'old_price': old_price, 'new_price': new_price,
                            'status': 'ok', 'msg': 'OK', 'sku': product.sku, 'url': mapping.url,
                            'product_id': product.id,  # ← nowe
                            'project_id': task.project_id,  # ← nowe
                            'my_price': product.my_price
                        }

                        if new_price is not None:
                            if old_price != new_price:
                                result_entry['status'] = 'change'

                            mapping.last_price = new_price
                            mapping.is_available = is_avail
                            mapping.last_checked_at = datetime.now()

                            history = PriceHistory(mapping_id=mapping.id, price=new_price, availability=is_avail)
                            db.session.add(history)

                            # --- AKTUALIZACJA CENY WŁASNEJ ---
                            if product.my_url and mapping.url.strip() == product.my_url.strip():
                                product.my_price = new_price
                                logger.info(f"    -> Zaktualizowano cenę własną produktu na: {new_price} PLN")
                        else:
                            result_entry['status'] = 'error'
                            result_entry['msg'] = 'Nie znaleziono ceny'
                        db.session.commit()

                        if new_price:
                            logger.info(f"Sukces: {new_price} PLN")
                        else:
                            logger.warning(f"Błąd: Nie znaleziono ceny")

                        scan_results.append(result_entry)

            logger.info("Zapisuję wyniki do bazy danych...")
            task.last_run_date = today_date
            db.session.commit()

            if scan_results:
                try:
                    logger.info(f"Generuję raport i wysyłam maila ({len(scan_results)} produktów)...")
                    send_enhanced_report(task_name, scan_results)
                    logger.info(f"[SCHEDULER] Zadanie wykonane. Raport wysłany!")
                except Exception as e:
                    logger.error(f"[SCHEDULER ERROR] Zadanie wykonane, ale błąd wysyłki: {e}", exc_info=True)
        close_batch_session(batch_session)


@scheduler.task('cron', id='sote_sales_sync', hour=3, minute=0)
def sync_sote_sales_daily():
    with app.app_context():
        logger.info("--- [SOTE SYNC] Rozpoczynam nocną synchronizację sprzedaży ---")

        projects = Project.query.filter_by(api_type='SOTE').all()
        today = date.today()
        yesterday = today - timedelta(days=1)

        for proj in projects:
            if not proj.api_url or not proj.api_user or not proj.api_password:
                logger.warning(f"[SOTE SYNC] Projekt {proj.name} nie ma pełnych danych API. Pomijam.")
                continue

            logger.info(f"[SOTE SYNC] Sprawdzanie historii dla projektu: {proj.name}")

            # 1. Sprawdzamy na jakiej dacie zatrzymała się baza tego konkretnego projektu
            last_record = SalesHistory.query.join(Product).filter(Product.project_id == proj.id).order_by(
                SalesHistory.date.desc()).first()

            if last_record:
                # Zaczynamy nadrabianie od dnia po ostatnim wpisie
                current_date = last_record.date + timedelta(days=1)
            else:
                # Jeśli baza jest całkowicie pusta, zasysamy z 7 dni wstecz na dobry początek
                current_date = today - timedelta(days=7)

            # Jeśli baza jest spójna i nie ma braków
            if current_date > yesterday:
                logger.info(f"[SOTE SYNC] Projekt {proj.name} jest aktualny do wczoraj. Brak braków.")
                continue

            active_products = Product.query.filter_by(project_id=proj.id, is_active=True).all()
            days_processed = 0

            # 2. PĘTLA: Pobieramy dane dzień po dniu aż do 'wczoraj'
            while current_date <= yesterday:
                logger.info(f"[SOTE SYNC] {proj.name} -> Zaciąganie danych dla: {current_date}")

                sales_data = fetch_sales_for_date(proj.api_url, proj.api_user, proj.api_password, current_date)

                for p in active_products:
                    sku_upper = str(p.sku).strip().upper() if p.sku else None
                    if not sku_upper: continue

                    qty = int(sales_data.get(sku_upper, {}).get('qty', 0))
                    revenue = float(sales_data.get(sku_upper, {}).get('revenue', 0.0))

                    # Wrzucamy dane, chroniąc się przed duplikatami
                    existing = SalesHistory.query.filter_by(product_id=p.id, date=current_date).first()
                    if existing:
                        existing.quantity = qty
                        existing.revenue = revenue
                    else:
                        new_history = SalesHistory(product_id=p.id, date=current_date, quantity=qty, revenue=revenue)
                        db.session.add(new_history)

                db.session.commit()
                current_date += timedelta(days=1)
                days_processed += 1

            logger.info(f"[SOTE SYNC] Zakończono projekt {proj.name}. Nadrobiono {days_processed} dni.")

# --- RĘCZNE WYMUSZENIE SKANOWANIA ---
@app.route('/project/<int:project_id>/scheduler/run-all', methods=['POST'])
@login_required
def run_all_tasks(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        return redirect(url_for('projects'))

    tasks = ScheduledTask.query.filter_by(project_id=project.id, is_active=True).all()

    if not tasks:
        flash('Brak aktywnych zadań.', category='warning')
        return redirect(url_for('project_scheduler', project_id=project.id))

    logger.info(f"--- [FORCE RUN] Start {len(tasks)} zadań ---")
    total_scanned = 0
    batch_session = init_batch_session()

    for task in tasks:
        task_label = f"Raport: Marka {task.brand.name}" if task.brand else "Raport: Cały Projekt"
        task_name = f"{task_label} (Wymuszony)"

        scan_results = []
        query = Product.query.filter_by(project_id=task.project_id, is_active=True)
        if task.brand_id:
            query = query.filter_by(brand_id=task.brand_id)
        products = query.all()

        for product in products:
            logger.info(f"[SCAN PRODUKT] {product.title} (SKU: {product.sku})")
            for mapping in product.mappings:
                if mapping.is_active:
                    old_price = mapping.last_price

                    try:
                        result = get_current_price(mapping.url, batch_session)
                        if result and isinstance(result, tuple):
                            new_price, is_avail = result
                        else:
                            new_price = None
                            is_avail = False
                    except Exception as e:
                        logger.warning(f"Error scanning {mapping.url}: {e}")
                        new_price = None
                        is_avail = False

                    result_entry = {
                        'product': product.title, 'shop': mapping.shop.name,
                        'old_price': old_price, 'new_price': new_price,
                        'status': 'ok', 'msg': 'OK', 'sku': product.sku, 'url': mapping.url,
                        'product_id': product.id,  # ← nowe
                        'project_id': task.project_id,  # ← nowe
                        'my_price': product.my_price
                    }

                    if new_price is not None:
                        if old_price != new_price:
                            result_entry['status'] = 'change'

                        mapping.last_price = new_price
                        mapping.is_available = is_avail
                        mapping.last_checked_at = datetime.now()

                        history = PriceHistory(mapping_id=mapping.id, price=new_price, availability=is_avail)
                        db.session.add(history)

                        # --- AKTUALIZACJA CENY WŁASNEJ ---
                        if product.my_url and mapping.url.strip() == product.my_url.strip():
                            product.my_price = new_price
                            logger.info(f"    -> Zaktualizowano cenę własną produktu na: {new_price} PLN")
                    else:
                        result_entry['status'] = 'error'
                        result_entry['msg'] = 'Nie znaleziono ceny'
                    db.session.commit()

                    scan_results.append(result_entry)
                    total_scanned += 1

        task.last_run_date = date.today()
        db.session.commit()

        if scan_results:
            try:
                send_enhanced_report(task_name, scan_results)
            except Exception as e:
                logger.error(f"Error sending report: {e}", exc_info=True)
    close_batch_session(batch_session)
    flash(f'Zakończono. Sprawdzono {total_scanned} linków.', category='success')
    return redirect(url_for('project_scheduler', project_id=project.id))


# --- URUCHOMIENIE POJEDYNCZEGO ZADANIA ---
@app.route('/project/<int:project_id>/scheduler/<int:task_id>/run', methods=['POST'])
@login_required
def run_single_task(project_id, task_id):
    task = ScheduledTask.query.get_or_404(task_id)
    if task.project_id != project_id:
        return redirect(url_for('projects'))

    task_label = f"Raport: Marka {task.brand.name}" if task.brand else "Raport: Cały Projekt"
    task_name = f"{task_label} (Pojedynczy)"

    scan_results = []
    query = Product.query.filter_by(project_id=task.project_id, is_active=True)
    if task.brand_id:
        query = query.filter_by(brand_id=task.brand_id)
    products = query.all()

    batch_session = init_batch_session()

    for product in products:
        logger.info(f"[SCAN PRODUKT] {product.title} (SKU: {product.sku})")
        for mapping in product.mappings:
            if mapping.is_active:
                logger.info(f"  -> [{mapping.shop.name}] {mapping.url}")
                old_price = mapping.last_price

                try:
                    result = get_current_price(mapping.url, batch_session)
                    if result and isinstance(result, tuple):
                        new_price, is_avail = result
                        logger.info(f"  -> WYNIK: {new_price} PLN | dostępny: {is_avail}")
                    else:
                        new_price = None
                        is_avail = False
                        logger.warning(f"  -> WYNIK: brak ceny")
                except Exception as e:
                    logger.warning(f"Error scanning {mapping.url}: {e}")
                    new_price = None
                    is_avail = False

                result_entry = {
                    'product': product.title, 'shop': mapping.shop.name,
                    'old_price': old_price, 'new_price': new_price,
                    'status': 'ok', 'msg': 'OK', 'sku': product.sku, 'url': mapping.url,
                    'product_id': product.id,  # ← nowe
                    'project_id': task.project_id,  # ← nowe
                    'my_price': product.my_price
                }

                if new_price is not None:
                    if old_price != new_price:
                        result_entry['status'] = 'change'

                    mapping.last_price = new_price
                    mapping.is_available = is_avail
                    mapping.last_checked_at = datetime.now()

                    history = PriceHistory(mapping_id=mapping.id, price=new_price, availability=is_avail)
                    db.session.add(history)

                    # --- AKTUALIZACJA CENY WŁASNEJ ---
                    if product.my_url and mapping.url.strip() == product.my_url.strip():
                        product.my_price = new_price
                        logger.info(f"    -> Zaktualizowano cenę własną produktu na: {new_price} PLN")
                else:
                    result_entry['status'] = 'error'
                    result_entry['msg'] = 'Nie znaleziono ceny'
                db.session.commit()

                scan_results.append(result_entry)

    task.last_run_date = date.today()
    db.session.commit()

    if scan_results:
        try:
            send_enhanced_report(task_name, scan_results)
            flash(f'Zadanie wykonane. Raport wysłany!', category='success')
        except Exception as e:
            logger.error(f"Error sending report: {e}", exc_info=True)
            flash(f'Zadanie wykonane, błąd wysyłki.', category='warning')

    close_batch_session(batch_session)
    return redirect(url_for('project_scheduler', project_id=project_id))


# Wyświetlanie strony raportów
@app.route('/project/<int:project_id>/reports')
@login_required
def project_reports(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    brands = Brand.query.join(Product).filter(Product.project_id == project.id).distinct().all()

    today = date.today()
    start_date = today.replace(day=1)

    return render_template('reports.html', project=project, brands=brands, today=today, start_date=start_date)


# Generowanie CSV
@app.route('/project/<int:project_id>/reports/download', methods=['POST'])
@login_required
def download_report(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    brand_id = request.form.get('brand_id')

    query = db.session.query(PriceHistory).join(ProductMapping).join(Product).join(Shop)

    query = query.filter(Product.project_id == project.id)

    if start_date_str:
        query = query.filter(PriceHistory.scraped_at >= f"{start_date_str} 00:00:00")
    if end_date_str:
        query = query.filter(PriceHistory.scraped_at <= f"{end_date_str} 23:59:59")

    if brand_id and brand_id != 'all':
        query = query.filter(Product.brand_id == int(brand_id))

    history_data = query.order_by(PriceHistory.scraped_at.desc(), Product.title).all()

    si = io.StringIO()
    cw = csv.writer(si, delimiter=';')

    cw.writerow(['Data', 'Godzina', 'SKU', 'Produkt', 'Marka', 'Sklep', 'Cena (PLN)', 'Dostępność', 'Link'])

    for row in history_data:
        date_str = row.scraped_at.strftime('%Y-%m-%d')
        time_str = row.scraped_at.strftime('%H:%M')

        cw.writerow([
            date_str,
            time_str,
            row.mapping.product.sku,
            row.mapping.product.title,
            row.mapping.product.brand.name if row.mapping.product.brand else 'Brak',
            row.mapping.shop.name,
            str(row.price).replace('.', ','),
            'Dostępny' if row.availability else 'Niedostępny',
            row.mapping.url
        ])

    output = make_response(si.getvalue().encode('utf-8-sig'))
    output.headers["Content-Disposition"] = f"attachment; filename=raport_cen_{start_date_str}_{end_date_str}.csv"
    output.headers["Content-type"] = "text/csv"

    return output


@app.route('/project/<int:project_id>/analysis')
@login_required
def project_analysis(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    products = Product.query.filter_by(project_id=project.id).all()

    competitor_stats = {}
    global_position = {'cheaper': 0, 'expensive': 0, 'equal': 0}

    opportunities = []
    threats = []

    for p in products:
        if not p.my_price: continue

        active_mappings = []
        for m in p.mappings:

            if not m.is_active or not m.last_price:
                continue

            if p.my_url and m.url.strip() == p.my_url.strip():
                continue

            active_mappings.append(m)

        if not active_mappings: continue

        market_prices = [m.last_price for m in active_mappings]
        min_market_price = min(market_prices)

        if p.my_price < min_market_price:
            global_position['cheaper'] += 1
            diff_percent = ((min_market_price - p.my_price) / min_market_price) * 100
            if diff_percent > 15:
                opportunities.append({
                    'product': p,
                    'diff': round(diff_percent, 1),
                    'market_price': min_market_price
                })

        elif p.my_price > min_market_price:
            global_position['expensive'] += 1
            diff_percent = ((p.my_price - min_market_price) / min_market_price) * 100
            if diff_percent > 10:
                threats.append({
                    'product': p,
                    'diff': round(diff_percent, 1),
                    'market_price': min_market_price
                })
        else:
            global_position['equal'] += 1

        for m in active_mappings:
            shop_name = m.shop.name
            if shop_name not in competitor_stats:
                competitor_stats[shop_name] = {'id': m.shop.id, 'name': shop_name, 'wins': 0, 'losses': 0, 'draws': 0,
                                               'total': 0}

            stats = competitor_stats[shop_name]
            stats['total'] += 1

            if p.my_price < m.last_price:
                stats['wins'] += 1
            elif p.my_price > m.last_price:
                stats['losses'] += 1
            else:
                stats['draws'] += 1

    sorted_competitors = sorted(competitor_stats.values(), key=lambda x: x['losses'], reverse=True)

    opportunities = sorted(opportunities, key=lambda x: x['diff'], reverse=True)[:5]  # Top 5
    threats = sorted(threats, key=lambda x: x['diff'], reverse=True)[:5]  # Top 5

    return render_template('analysis.html',
                           project=project,
                           competitors=sorted_competitors,
                           global_position=global_position,
                           opportunities=opportunities,
                           threats=threats)

@app.route('/project/<int:project_id>/competitors')
@login_required
def competitors_list(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    search_query = request.args.get('q', '')
    sort_by = request.args.get('sort', 'shared_count')
    sort_order = request.args.get('order', 'desc')
    page = request.args.get('page', 1, type=int)
    view_mode = request.args.get('view', 'grid') # Tryb wyświetlania (grid/list)

    from sqlalchemy import func, case, or_

    pi_expr = (ProductMapping.last_price / Product.my_price) * 100
    shared_count_expr = func.count(Product.id)
    avg_pi_expr = func.avg(pi_expr)
    cheaper_count_expr = func.sum(case((ProductMapping.last_price < Product.my_price, 1), else_=0))

    query = db.session.query(
        Shop,
        shared_count_expr.label('shared_count'),
        avg_pi_expr.label('avg_pi'),
        cheaper_count_expr.label('cheaper_count')
    ).select_from(ProductMapping).join(Product).join(Shop).filter(
        Product.project_id == project.id,
        Product.is_active == True,
        Product.my_price > 0,
        ProductMapping.is_active == True,
        ProductMapping.is_available == True,
        ProductMapping.last_price > 0,
        or_(Product.my_url == None, ProductMapping.url != Product.my_url),
        ProductMapping.last_price >= Product.my_price * 0.2,
        ProductMapping.last_price <= Product.my_price * 5
    )

    if search_query:
        query = query.filter(
            or_(Shop.name.ilike(f'%{search_query}%'), Shop.domain.ilike(f'%{search_query}%'))
        )

    query = query.group_by(Shop.id)

    if sort_by == 'name':
        query = query.order_by(Shop.name.desc() if sort_order == 'desc' else Shop.name.asc())
    elif sort_by == 'shared_count':
        query = query.order_by(shared_count_expr.desc() if sort_order == 'desc' else shared_count_expr.asc())
    elif sort_by == 'avg_pi':
        query = query.order_by(avg_pi_expr.desc() if sort_order == 'desc' else avg_pi_expr.asc())
    elif sort_by == 'cheaper_count':
        query = query.order_by(cheaper_count_expr.desc() if sort_order == 'desc' else cheaper_count_expr.asc())

    pagination = query.paginate(page=page, per_page=20, error_out=False)
    competitor_stats = pagination.items

    return render_template('competitors_list.html',
                           project=project,
                           competitor_stats=competitor_stats,
                           pagination=pagination,
                           current_filters={'q': search_query, 'sort': sort_by, 'order': sort_order, 'view': view_mode})

@app.route('/project/<int:project_id>/margin')
@login_required
def project_margin(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    search_query = request.args.get('q', '')
    brand_filter = request.args.get('brand', '')
    sort_by = request.args.get('sort', 'margin_pct')
    sort_order = request.args.get('order', 'asc')
    page = request.args.get('page', 1, type=int)

    query = Product.query.filter(
        Product.project_id == project.id,
        Product.is_active == True,
        Product.purchase_price != None,
        Product.my_price != None,
        Product.my_price > 0
    )

    if search_query:
        query = query.filter(
            (Product.title.ilike(f'%{search_query}%')) |
            (Product.sku.ilike(f'%{search_query}%'))
        )

    if brand_filter and brand_filter.isdigit():
        query = query.filter_by(brand_id=int(brand_filter))

    # --- UWAGA: TUTAJ COFAMY WCIĘCIE O JEDEN POZIOM ---
    # --- Sortowanie na poziomie Bazy Danych ---
    margin_expr = Product.my_price - Product.purchase_price
    margin_pct_expr = (Product.my_price - Product.purchase_price) / Product.my_price

    if sort_by == 'margin_pln':
        query = query.order_by(margin_expr.desc() if sort_order == 'desc' else margin_expr.asc())
    elif sort_by == 'margin_pct':
        query = query.order_by(margin_pct_expr.desc() if sort_order == 'desc' else margin_pct_expr.asc())
    elif sort_by == 'title':
        query = query.order_by(Product.title.desc() if sort_order == 'desc' else Product.title.asc())
    elif sort_by == 'price_index':
        from sqlalchemy import func, or_
        avg_comp_stmt = db.session.query(func.avg(ProductMapping.last_price)).filter(
            ProductMapping.product_id == Product.id,
            ProductMapping.is_active == True,
            ProductMapping.is_available == True,
            ProductMapping.last_price > 0,
            or_(Product.my_url == None, ProductMapping.url != Product.my_url),
            ProductMapping.last_price >= Product.my_price * 0.2,
            ProductMapping.last_price <= Product.my_price * 5
        ).correlate(Product).scalar_subquery()

        pi_expr = (avg_comp_stmt / Product.my_price) * 100

        if sort_order == 'desc':
            query = query.order_by(pi_expr.desc())
        else:
            query = query.order_by(pi_expr.is_(None), pi_expr.asc())

    from sqlalchemy import func as sqlfunc

    margin_expr = Product.my_price - Product.purchase_price
    margin_pct_expr = (Product.my_price - Product.purchase_price) / Product.my_price * 100

    agg = query.with_entities(
        sqlfunc.avg(margin_pct_expr).label('avg_pct'),
        sqlfunc.avg(margin_expr).label('avg_pln'),
        sqlfunc.count(Product.id).filter(margin_pct_expr < 10).label('below_threshold'),
    ).first()

    summary = {
        'avg_pct': round(agg.avg_pct, 1) if agg.avg_pct else 0,
        'avg_pln': round(agg.avg_pln, 2) if agg.avg_pln else 0,
        'below_threshold': agg.below_threshold or 0,
    }

    # --- Paginacja ---
    pagination = query.paginate(page=page, per_page=20, error_out=False)
    products = pagination.items

    analyzed_products = []
    for p in products:
        margin_pln = p.my_price - p.purchase_price
        margin_pct = (margin_pln / p.my_price * 100)

        min_market_price = None
        active_mappings = [m for m in p.mappings if
                           m.is_active and m.last_price and (not p.my_url or m.url.strip() != p.my_url.strip())]
        if active_mappings:
            min_market_price = min([m.last_price for m in active_mappings])

        price_index = None
        valid_prices = [
            m.last_price for m in active_mappings
            if m.is_available and m.last_price > 0
               and (p.my_price * 0.2 <= m.last_price <= p.my_price * 5)
        ]
        if valid_prices:
            avg_comp = sum(valid_prices) / len(valid_prices)
            price_index = round((avg_comp / p.my_price) * 100, 1)

        analyzed_products.append({
            'product': p,
            'margin_pln': margin_pln,
            'margin_pct': margin_pct,
            'min_market_price': min_market_price,
            'price_index': price_index
        })

    available_brands = db.session.query(Brand).join(Product).filter(
        Product.project_id == project.id,
        Product.is_active == True,
        Product.purchase_price != None
    ).distinct().order_by(Brand.name).all()

    return render_template('margin_analysis.html',
                           project=project,
                           analyzed_products=analyzed_products,
                           pagination=pagination,
                           available_brands=available_brands,
                           summary=summary,
                           current_filters={
                               'q': search_query,
                               'brand': brand_filter,
                               'sort': sort_by,
                               'order': sort_order
                           })


@app.route('/project/<int:project_id>/sales-report')
@login_required
def sales_report(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    from sqlalchemy import func, case
    from datetime import date

    today = date.today()
    first_day = today.replace(day=1)

    start_date_str = request.args.get('start_date', first_day.strftime('%Y-%m-%d'))
    end_date_str = request.args.get('end_date', today.strftime('%Y-%m-%d'))
    brand_filter = request.args.get('brand', '')
    sort_by = request.args.get('sort', 'revenue')
    sort_order = request.args.get('order', 'desc')
    page = request.args.get('page', 1, type=int)

    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        start_date, end_date = first_day, today

    # Definicje wyrażeń SQL dla agregatów
    qty_expr = func.sum(SalesHistory.quantity)
    rev_expr = func.sum(SalesHistory.revenue)
    # Zysk = Przychód - (Cena Zakupu * Ilość)
    profit_expr = rev_expr - (func.coalesce(Product.purchase_price, 0) * qty_expr)
    # Marża % = (Zysk / Przychód) * 100
    margin_pct_expr = case((rev_expr > 0, (profit_expr / rev_expr) * 100), else_=0)

    query = db.session.query(
        Product,
        qty_expr.label('total_qty'),
        rev_expr.label('total_rev'),
        profit_expr.label('total_profit'),
        margin_pct_expr.label('margin_pct')
    ).join(SalesHistory).filter(
        Product.project_id == project.id,
        SalesHistory.date >= start_date,
        SalesHistory.date <= end_date,
        SalesHistory.quantity > 0
    )

    if brand_filter and brand_filter.isdigit():
        query = query.filter(Product.brand_id == int(brand_filter))

    query = query.group_by(Product.id)

    # Sortowanie serwerowe
    if sort_by == 'title':
        query = query.order_by(Product.title.desc() if sort_order == 'desc' else Product.title.asc())
    elif sort_by == 'brand':
        query = query.join(Brand, isouter=True).order_by(
            Brand.name.desc() if sort_order == 'desc' else Brand.name.asc())
    elif sort_by == 'qty':
        query = query.order_by(qty_expr.desc() if sort_order == 'desc' else qty_expr.asc())
    elif sort_by == 'revenue':
        query = query.order_by(rev_expr.desc() if sort_order == 'desc' else rev_expr.asc())
    elif sort_by == 'profit':
        query = query.order_by(profit_expr.desc() if sort_order == 'desc' else profit_expr.asc())
    elif sort_by == 'margin':
        query = query.order_by(margin_pct_expr.desc() if sort_order == 'desc' else margin_pct_expr.asc())

    pagination = query.paginate(page=page, per_page=50, error_out=False)

    # Globalne statystyki dla kafelków
    global_stats = db.session.query(func.sum(SalesHistory.quantity), func.sum(SalesHistory.revenue)).join(
        Product).filter(
        Product.project_id == project.id, SalesHistory.date >= start_date, SalesHistory.date <= end_date
    )
    if brand_filter and brand_filter.isdigit():
        global_stats = global_stats.filter(Product.brand_id == int(brand_filter))

    g_qty, g_rev = global_stats.first()
    available_brands = db.session.query(Brand).join(Product).filter(
        Product.project_id == project.id).distinct().order_by(Brand.name).all()

    return render_template('sales_report.html',
                           project=project,
                           pagination=pagination,
                           available_brands=available_brands,
                           g_qty=g_qty or 0,
                           g_rev=g_rev or 0,
                           current_filters={
                               'start_date': start_date_str, 'end_date': end_date_str,
                               'brand': brand_filter, 'sort': sort_by, 'order': sort_order
                           })

@app.route('/project/<int:project_id>/margin-by-brand')
@login_required
def project_margin_by_brand(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    page = request.args.get('page', 1, type=int)
    sort_by = request.args.get('sort', 'avg_pct')
    sort_order = request.args.get('order', 'desc')

    from sqlalchemy import func, case, or_

    margin_expr = Product.my_price - Product.purchase_price
    margin_pct_expr = (Product.my_price - Product.purchase_price) / Product.my_price * 100

    total_products_expr = func.count(Product.id)
    avg_pct_expr = func.avg(margin_pct_expr)
    avg_pln_expr = func.avg(margin_expr)
    below_threshold_expr = func.sum(case((margin_pct_expr < 10, 1), else_=0))

    # --- SQL dla Price Index Marek ---
    avg_comp_stmt = db.session.query(func.avg(ProductMapping.last_price)).filter(
        ProductMapping.product_id == Product.id,
        ProductMapping.is_active == True,
        ProductMapping.is_available == True,
        ProductMapping.last_price > 0,
        or_(Product.my_url == None, ProductMapping.url != Product.my_url),
        ProductMapping.last_price >= Product.my_price * 0.2,
        ProductMapping.last_price <= Product.my_price * 5
    ).correlate(Product).scalar_subquery()

    # Wyliczamy PI dla produktu, a main query wyliczy z tego średnią (func.avg)
    avg_pi_expr = func.avg((avg_comp_stmt / Product.my_price) * 100)

# ... (wcześniejsza część funkcji zostaje bez zmian)

    # Dodajemy Brand.id do select_from i group_by
    query = db.session.query(
        Brand.id.label('brand_id'),
        Brand.name.label('brand_name'),
        total_products_expr.label('total_products'),
        avg_pct_expr.label('avg_pct'),
        avg_pln_expr.label('avg_pln'),
        below_threshold_expr.label('below_threshold'),
        avg_pi_expr.label('avg_pi')
    ).select_from(Product).join(Brand).filter(
        Product.project_id == project.id,
        Product.is_active == True,
        Product.purchase_price != None,
        Product.my_price != None,
        Product.my_price > 0
    ).group_by(Brand.id, Brand.name)

    # Logika sortowania (zostaje bez zmian)
    if sort_by == 'brand_name':
        query = query.order_by(Brand.name.desc() if sort_order == 'desc' else Brand.name.asc())
    elif sort_by == 'total_products':
        query = query.order_by(total_products_expr.desc() if sort_order == 'desc' else total_products_expr.asc())
    elif sort_by == 'avg_pln':
        query = query.order_by(avg_pln_expr.desc() if sort_order == 'desc' else avg_pln_expr.asc())
    elif sort_by == 'below_threshold':
        query = query.order_by(below_threshold_expr.desc() if sort_order == 'desc' else below_threshold_expr.asc())
    elif sort_by == 'avg_pi':
        if sort_order == 'desc':
            query = query.order_by(avg_pi_expr.desc())
        else:
            query = query.order_by(avg_pi_expr.is_(None), avg_pi_expr.asc())
    else:
        query = query.order_by(avg_pct_expr.desc() if sort_order == 'desc' else avg_pct_expr.asc())

    pagination = query.paginate(page=page, per_page=20, error_out=False)
    brand_stats_raw = pagination.items

    brand_stats = []
    for stat in brand_stats_raw:
        brand_stats.append({
            'brand_id': stat.brand_id, # Wyciągamy wygenerowane ID
            'brand_name': stat.brand_name,
            'total_products': stat.total_products,
            'avg_pct': stat.avg_pct,
            'avg_pln': stat.avg_pln,
            'below_threshold': stat.below_threshold,
            'avg_pi': round(stat.avg_pi, 1) if stat.avg_pi else None
        })

    return render_template('margin_by_brand.html',
                           project=project,
                           brand_stats=brand_stats,
                           pagination=pagination,
                           current_filters={'sort': sort_by, 'order': sort_order})


@app.route('/project/<int:project_id>/brand-monitor')
@login_required
def brand_monitoring(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    brand_id = request.args.get('brand', type=int)
    page = request.args.get('page', 1, type=int)
    sort_by = request.args.get('sort', 'price_index')
    sort_order = request.args.get('order', 'asc')

    available_brands = db.session.query(Brand).join(Product).filter(
        Product.project_id == project.id,
        Product.is_active == True
    ).distinct().order_by(Brand.name).all()

    brand = None
    brand_stats = {}
    pagination = None
    brand_data = []

    if brand_id:
        brand = Brand.query.get_or_404(brand_id)

        from sqlalchemy import func, case, or_


        # Wyrażenie do wyliczenia Price Indexu w locie przez bazę danych
        avg_comp_stmt = db.session.query(func.avg(ProductMapping.last_price)).filter(
            ProductMapping.product_id == Product.id,
            ProductMapping.is_active == True,
            ProductMapping.is_available == True,  # Tylko dostępne!
            ProductMapping.last_price > 0,
            or_(Product.my_url == None, ProductMapping.url != Product.my_url),
            ProductMapping.last_price >= Product.my_price * 0.2,  # Zabezpieczenie przed anomaliami
            ProductMapping.last_price <= Product.my_price * 5
        ).correlate(Product).scalar_subquery()

        pi_expr = (avg_comp_stmt / Product.my_price) * 100

        # --- 1. STATYSTYKI GŁÓWNE ---
        stats_query = db.session.query(
            func.count(Product.id).label('total'),
            func.avg(pi_expr).label('avg_pi'),
            func.sum(case((pi_expr < 100, 1), else_=0)).label('conflicts'),
            func.sum(case((pi_expr > 115, 1), else_=0)).label('opportunities')
        ).filter(
            Product.project_id == project.id,
            Product.brand_id == brand.id,
            Product.is_active == True,
            Product.my_price > 0
        ).first()

        brand_stats = {
            'total': stats_query.total or 0,
            'avg_pi': round(stats_query.avg_pi, 1) if stats_query.avg_pi else 0,
            'conflicts': stats_query.conflicts or 0,
            'opportunities': stats_query.opportunities or 0
        }

        # --- 2. ZAPYTANIE O PRODUKTY I SORTOWANIE ---
        query = Product.query.filter_by(
            project_id=project.id,
            brand_id=brand.id,
            is_active=True
        )

        # Logika sortowania z zabezpieczeniem (wypycha wartości NULL na sam dół)
        if sort_by == 'title':
            query = query.order_by(Product.title.desc() if sort_order == 'desc' else Product.title.asc())
        elif sort_by == 'price':
            query = query.order_by(Product.my_price.is_(None),
                                   Product.my_price.desc() if sort_order == 'desc' else Product.my_price.asc())
        elif sort_by == 'price_index':
            query = query.order_by(pi_expr.is_(None), pi_expr.desc() if sort_order == 'desc' else pi_expr.asc())

        pagination = query.paginate(page=page, per_page=20, error_out=False)

        for p in pagination.items:
            sorted_mappings = sorted(p.mappings, key=lambda m: (
                m.last_price if m.last_price else float('inf'),
                m.shop.name
            ))

            # Ścisła synchronizacja z SQL, żeby frontend pokazywał to samo co sortuje baza
            valid_prices = [
                m.last_price for m in sorted_mappings
                if m.is_active and m.is_available and m.last_price and m.last_price > 0
                   and (not p.my_url or m.url != p.my_url)
                   and (p.my_price * 0.2 <= m.last_price <= p.my_price * 5)
            ]

            row_pi = None
            if valid_prices and p.my_price:
                row_pi = round(((sum(valid_prices) / len(valid_prices)) / p.my_price) * 100, 1)

            brand_data.append({
                'product': p,
                'mappings': sorted_mappings,
                'price_index': row_pi
            })

    return render_template('brand_monitoring.html',
                           project=project,
                           available_brands=available_brands,
                           brand=brand,
                           brand_stats=brand_stats,
                           brand_data=brand_data,
                           pagination=pagination,
                           current_filters={'brand': brand_id, 'sort': sort_by, 'order': sort_order})

@app.route('/project/<int:project_id>/competitor/<int:shop_id>')
@login_required
def competitor_analysis(project_id, shop_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    shop = Shop.query.get_or_404(shop_id)

    # Szukamy aktywnych powiązań tego sklepu w tym projekcie
    mappings = ProductMapping.query.join(Product).filter(
        Product.project_id == project.id,
        Product.is_active == True,
        Product.my_price > 0,
        ProductMapping.shop_id == shop.id,
        ProductMapping.is_active == True,
        ProductMapping.is_available == True,
        ProductMapping.last_price > 0
    ).all()

    overlap_count = 0
    pi_list = []
    conflict_list = []  # Tzw. "Lista zapalna"

    for m in mappings:
        p = m.product
        # Filtrujemy anomalie i odrzucamy Twój własny sklep
        if (not p.my_url or m.url.strip() != p.my_url.strip()) and (p.my_price * 0.2 <= m.last_price <= p.my_price * 5):
            overlap_count += 1

            # Tutaj liczymy JEGO Price Index (Jego Cena / Twoja Cena * 100)
            pi = (m.last_price / p.my_price) * 100
            pi_list.append(pi)

            # Jeśli jego PI < 100, to znaczy, że on jest tańszy od Ciebie
            if pi < 100:
                conflict_list.append({
                    'product': p,
                    'his_price': m.last_price,
                    'my_price': p.my_price,
                    'pi': round(pi, 1),
                    'diff_pln': round(p.my_price - m.last_price, 2),
                    'url': m.url
                })

    avg_pi = round(sum(pi_list) / len(pi_list), 1) if pi_list else None

    # Sortujemy "listę zapalną" od produktów, gdzie podcina nas najmocniej (najniższe PI na górze)
    conflict_list = sorted(conflict_list, key=lambda x: x['pi'])

    return render_template('competitor_analysis.html',
                           project=project,
                           shop=shop,
                           overlap_count=overlap_count,
                           avg_pi=avg_pi,
                           conflict_list=conflict_list)


def send_async_email(app, msg):
    with app.app_context():
        try:
            mail.send(msg)
            logger.info("Powiadomienie o komentarzu wysłane.")
        except Exception as e:
            logger.error(f"Błąd wysyłania powiadomienia: {e}")


@app.route('/project/<int:project_id>/product/<int:product_id>/update-note', methods=['POST'])
@login_required
def update_strategic_note(project_id, product_id):
    product = Product.query.get_or_404(product_id)
    product.strategic_note = request.form.get('strategic_note')
    db.session.commit()
    flash('Zaktualizowano notatkę strategiczną.', 'success')
    return redirect(url_for('product_details', project_id=project_id, product_id=product_id))


@app.route('/project/<int:project_id>/product/<int:product_id>/add-comment', methods=['POST'])
@login_required
def add_product_comment(project_id, product_id):
    product = Product.query.get_or_404(product_id)
    content = request.form.get('content')

    if content and content.strip():
        comment = ProductComment(product_id=product.id, user_id=current_user.id, content=content.strip())
        db.session.add(comment)
        db.session.commit()

        # Przygotowanie i wysyłka maila w tle
        APP_URL = os.environ.get('APP_URL', 'http://127.0.0.1:5005')
        recipient = app.config.get('MAIL_RECIPIENT') or app.config['MAIL_DEFAULT_SENDER']
        msg = Message(f"Nowy komentarz do: {product.title}", recipients=[recipient])

        product_link = f"{APP_URL}/project/{project_id}/product/{product_id}"
        msg.html = f"""
        <div style="font-family: Arial, sans-serif; padding: 20px; background: #f8f9fa;">
            <div style="background: white; padding: 20px; border-radius: 8px; max-width: 600px; margin: 0 auto;">
                <h3 style="color: #0d6efd; margin-top: 0;">Nowy komentarz w systemie</h3>
                <p><strong>Użytkownik:</strong> {current_user.email}</p>
                <p><strong>Produkt:</strong> {product.title} (SKU: {product.sku})</p>
                <div style="background: #f1f3f5; padding: 15px; border-left: 4px solid #0d6efd; margin: 15px 0;">
                    {content.strip()}
                </div>
                <a href="{product_link}" style="background: #0d6efd; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; display: inline-block; margin-top: 10px;">Zobacz w aplikacji</a>
            </div>
        </div>
        """
        Thread(target=send_async_email, args=(app, msg)).start()

        flash('Komentarz dodany.', 'success')

    return redirect(url_for('product_details', project_id=project_id, product_id=product_id))


@app.route('/project/<int:project_id>/product/<int:product_id>/delete-comment/<int:comment_id>', methods=['POST'])
@login_required
def delete_product_comment(project_id, product_id, comment_id):
    comment = ProductComment.query.get_or_404(comment_id)
    db.session.delete(comment)
    db.session.commit()
    flash('Komentarz usunięty.', 'success')
    return redirect(url_for('product_details', project_id=project_id, product_id=product_id))


@app.route('/project/<int:project_id>/compare')
@login_required
def compare_projects(project_id):
    project1 = Project.query.get_or_404(project_id)
    if current_user not in project1.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    # Pobieramy inne projekty użytkownika do listy rozwijanej
    other_projects = [p for p in current_user.projects if p.id != project1.id]

    target_project_id = request.args.get('target', type=int)
    project2 = None

    stats = {}
    conflicts_price = []
    conflicts_avail = []
    only_in_p1 = []
    only_in_p2 = []

    if target_project_id:
        project2 = Project.query.get_or_404(target_project_id)
        if current_user not in project2.users:
            flash('Brak dostępu do projektu docelowego.', category='error')
            return redirect(url_for('project_dashboard', project_id=project1.id))

        # Pobieramy aktywne produkty z obu projektów
        p1_prods = Product.query.filter_by(project_id=project1.id, is_active=True).all()
        p2_prods = Product.query.filter_by(project_id=project2.id, is_active=True).all()

        # Funkcja normalizująca SKU (usuwa białe znaki, robi wielkie litery)
        def norm_sku(sku):
            return str(sku).strip().upper() if sku else None

        # Słowniki produktów (SKU -> Obiekt Produktu)
        p1_dict = {norm_sku(p.sku): p for p in p1_prods if norm_sku(p.sku)}
        p2_dict = {norm_sku(p.sku): p for p in p2_prods if norm_sku(p.sku)}

        # Zbiory unikalnych SKU
        p1_skus = set(p1_dict.keys())
        p2_skus = set(p2_dict.keys())

        # Magia Pythona: przecięcia i różnice zbiorów w locie
        common_skus = p1_skus.intersection(p2_skus)
        unique_p1_skus = p1_skus - p2_skus
        unique_p2_skus = p2_skus - p1_skus

        # Sprawdzamy konflikty tylko dla wspólnych produktów
        for sku in common_skus:
            prod1 = p1_dict[sku]
            prod2 = p2_dict[sku]

            # 1. Analiza cenowa
            if prod1.my_price != prod2.my_price:
                diff = (prod1.my_price or 0) - (prod2.my_price or 0)
                conflicts_price.append({
                    'sku': sku,
                    'p1_prod': prod1,
                    'p2_prod': prod2,
                    'diff': diff
                })

            # 2. Analiza magazynowa (ujednolicamy statusy, żeby wyłapać faktyczne różnice)
            def normalize_status(s):
                s_str = str(s).upper() if s else ''
                if any(x in s_str for x in ['IN STOCK', 'DOSTĘPNY', 'AVAILABLE']): return 'IN_STOCK'
                if any(x in s_str for x in ['OUT OF STOCK', 'NIEDOSTĘPNY']): return 'OUT_OF_STOCK'
                return 'OTHER'

            if normalize_status(prod1.availability) != normalize_status(prod2.availability):
                conflicts_avail.append({
                    'sku': sku,
                    'p1_prod': prod1,
                    'p2_prod': prod2
                })

        # Przepisujemy obiekty dla brakujących
        for sku in unique_p1_skus: only_in_p1.append(p1_dict[sku])
        for sku in unique_p2_skus: only_in_p2.append(p2_dict[sku])

        stats = {
            'common': len(common_skus),
            'price_conflicts': len(conflicts_price),
            'avail_conflicts': len(conflicts_avail),
            'unique_p1': len(unique_p1_skus),
            'unique_p2': len(unique_p2_skus)
        }

    return render_template('compare_projects.html',
                           project=project1,
                           other_projects=other_projects,
                           project2=project2,
                           stats=stats,
                           conflicts_price=conflicts_price,
                           conflicts_avail=conflicts_avail,
                           only_in_p1=only_in_p1,
                           only_in_p2=only_in_p2)

@app.route('/project/<int:project_id>/overview')
@login_required
def project_overview(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user not in project.users:
        flash('Brak dostępu.', category='error')
        return redirect(url_for('projects'))

    today = date.today()

    scans_today = db.session.query(PriceHistory).join(ProductMapping).join(Product).filter(
        Product.project_id == project.id,
        func.date(PriceHistory.scraped_at) == today
    ).count()

    errors_count = ProductMapping.query.join(Product).filter(
        Product.project_id == project.id,
        ProductMapping.is_active == True,
        (ProductMapping.last_price == None) | (ProductMapping.last_price == 0)
    ).count()

    recent_activity = db.session.query(PriceHistory).join(ProductMapping).join(Product).join(Shop).filter(
        Product.project_id == project.id
    ).order_by(PriceHistory.scraped_at.desc()).limit(10).all()

    last_scan_time = recent_activity[0].scraped_at if recent_activity else None

    # --- ETAP 3: GLOBALNY PRICE INDEX ---
    active_products = Product.query.filter_by(project_id=project.id, is_active=True).all()
    project_pi_list = []

    for p in active_products:
        if p.my_price and p.my_price > 0:
            valid_prices = [
                m.last_price for m in p.mappings
                if m.is_active and m.is_available and m.last_price and m.last_price > 0
                   and (not p.my_url or m.url.strip() != p.my_url.strip())
                   and (p.my_price * 0.2 <= m.last_price <= p.my_price * 5)
            ]
            if valid_prices:
                avg_comp = sum(valid_prices) / len(valid_prices)
                project_pi_list.append((avg_comp / p.my_price) * 100)

    global_pi = round(sum(project_pi_list) / len(project_pi_list), 1) if project_pi_list else None

    return render_template('overview.html',
                           project=project,
                           scans_today=scans_today,
                           errors_count=errors_count,
                           recent_activity=recent_activity,
                           last_scan_time=last_scan_time,
                           global_pi=global_pi)

# TWORZENIE ADMINA - inicjalizacja tylko przy pierwszym uruchomieniu
@app.route('/create-admin')
def create_admin():
    db.create_all()
    email = os.getenv('ADMIN_EMAIL')
    password = os.getenv('ADMIN_PASSWORD')

    if not email or not password:
        return "Błąd: Brak danych admina w pliku .env!"
    if not User.query.filter_by(email=email).first():
        hashed_pw = generate_password_hash(password, method='pbkdf2:sha256')
        new_user = User(email=email, password=hashed_pw)
        db.session.add(new_user)
        db.session.commit()
        return f"Stworzono admina ({email})! Teraz możesz się zalogować."
    return "Admin już istnieje."


# --- REJESTRACJA WIDOKÓW W FLASK-ADMIN ---
admin.add_view(UserModelView(User, db.session, name='Użytkownicy'))
admin.add_view(ProjectModelView(Project, db.session, name='Projekty'))
admin.add_view(MyModelView(Brand, db.session, name='Marki'))
admin.add_view(ProductModelView(Product, db.session, name='Produkty'))
admin.add_view(MyModelView(Shop, db.session, name='Sklepy'))
admin.add_view(MyModelView(ProductMapping, db.session, name='Linki (Mapping)'))
admin.add_view(MyModelView(ScheduledTask, db.session, name='Harmonogram'))
admin.add_view(MyModelView(PriceHistory, db.session, name='Historia Cen'))
admin.add_view(MyModelView(SalesHistory, db.session, name='Historia Sprzedaży'))

if __name__ == '__main__':
    with app.app_context():
        db.create_all()

    app.run(host='0.0.0.0', port=5005, debug=True, use_reloader=False)